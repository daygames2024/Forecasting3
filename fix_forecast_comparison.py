"""
Fixed forecast comparison code
Run with: python fix_forecast_comparison.py --old-forecast "forecast.csv" --actual-sales "sales.xlsx"
"""
import pandas as pd
from datetime import datetime
import argparse
import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl not available - will save CSV only (no Excel analysis report)")


def create_accuracy_analysis_excel(comparison_df, quarter_accuracy, excel_path, forecast_file, actuals_file, over_count, under_count):
    """
    Create detailed Excel workbook with multiple sheets analyzing forecast accuracy
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Summary
        ws_summary = wb.create_sheet('Summary', 0)
        ws_summary['A1'] = 'FORECAST ACCURACY ANALYSIS REPORT'
        ws_summary['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws_summary['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        summary_data = [
            ['', ''],
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Forecast File:', forecast_file],
            ['Actuals File:', actuals_file],
            ['', ''],
            ['OVERALL STATISTICS', ''],
            ['Total Comparisons:', len(comparison_df)],
            ['Average Error:', f"{comparison_df['Percentage_Error'].mean():.2f}%"],
            ['Median Error:', f"{comparison_df['Percentage_Error'].median():.2f}%"],
            ['Std Deviation:', f"{comparison_df['Percentage_Error'].std():.2f}%"],
            ['Max Over-Forecast:', f"{comparison_df['Percentage_Error'].max():.2f}%"],
            ['Max Under-Forecast:', f"{comparison_df['Percentage_Error'].min():.2f}%"],
            ['', ''],
            ['BIAS DETECTION', ''],
            ['Over-Forecasted Parts:', f"{over_count:,} ({over_count/len(comparison_df)*100:.1f}%)"],
            ['Under-Forecasted Parts:', f"{under_count:,} ({under_count/len(comparison_df)*100:.1f}%)"],
            ['', ''],
            ['ACCURACY BANDS', ''],
            ['Excellent (<10% error):', f"{(comparison_df['Percentage_Error'].abs() < 10).sum():,} ({(comparison_df['Percentage_Error'].abs() < 10).sum()/len(comparison_df)*100:.1f}%)"],
            ['Good (10-25% error):', f"{((comparison_df['Percentage_Error'].abs() >= 10) & (comparison_df['Percentage_Error'].abs() < 25)).sum():,} ({((comparison_df['Percentage_Error'].abs() >= 10) & (comparison_df['Percentage_Error'].abs() < 25)).sum()/len(comparison_df)*100:.1f}%)"],
            ['Fair (25-50% error):', f"{((comparison_df['Percentage_Error'].abs() >= 25) & (comparison_df['Percentage_Error'].abs() < 50)).sum():,} ({((comparison_df['Percentage_Error'].abs() >= 25) & (comparison_df['Percentage_Error'].abs() < 50)).sum()/len(comparison_df)*100:.1f}%)"],
            ['Poor (50-100% error):', f"{((comparison_df['Percentage_Error'].abs() >= 50) & (comparison_df['Percentage_Error'].abs() < 100)).sum():,} ({((comparison_df['Percentage_Error'].abs() >= 50) & (comparison_df['Percentage_Error'].abs() < 100)).sum()/len(comparison_df)*100:.1f}%)"],
            ['Critical (>100% error):', f"{(comparison_df['Percentage_Error'].abs() >= 100).sum():,} ({(comparison_df['Percentage_Error'].abs() >= 100).sum()/len(comparison_df)*100:.1f}%)"],
        ]

        for r_idx, row_data in enumerate(summary_data, 1):
            ws_summary[f'A{r_idx}'] = row_data[0]
            ws_summary[f'B{r_idx}'] = row_data[1]
            if row_data[0].isupper() and row_data[0]:
                ws_summary[f'A{r_idx}'].font = Font(bold=True, color='FFFFFF')
                ws_summary[f'A{r_idx}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40

        # Sheet 2: By Quarter
        ws_quarter = wb.create_sheet('By_Quarter', 1)
        ws_quarter['A1'] = 'Accuracy by Quarter'
        ws_quarter['A1'].font = Font(bold=True, size=14)

        quarter_df = quarter_accuracy.reset_index()
        quarter_df.columns = ['Quarter', 'Mean Error %', 'Median Error %', 'Std Dev %']

        for r_idx, row in enumerate(dataframe_to_rows(quarter_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_quarter.cell(row=r_idx+1, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        ws_quarter.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D']:
            ws_quarter.column_dimensions[col].width = 18

        # Sheet 3: Top Errors
        ws_errors = wb.create_sheet('Top_Errors', 2)
        ws_errors['A1'] = 'Top 50 Biggest Forecast Errors'
        ws_errors['A1'].font = Font(bold=True, size=14)

        top_50_errors = comparison_df.nlargest(50, 'Absolute_Error')[['Part_number', 'Year_Q', 'Forecast', 'Actual', 'Error', 'Percentage_Error']]

        for r_idx, row in enumerate(dataframe_to_rows(top_50_errors, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_errors.cell(row=r_idx+1, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                # Color code errors
                elif c_idx == 6 and isinstance(value, (int, float)):  # Percentage_Error column
                    if value > 100:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
                    elif value > 50:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow

        ws_errors.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_errors.column_dimensions[col].width = 15

        # Sheet 4: All Parts
        ws_all = wb.create_sheet('All_Parts', 3)
        ws_all['A1'] = 'Complete Forecast vs Actual Comparison'
        ws_all['A1'].font = Font(bold=True, size=14)

        all_data = comparison_df[['Part_number', 'Year_Q', 'Forecast', 'Actual', 'Error', 'Percentage_Error']].sort_values('Percentage_Error', key=abs, ascending=False)

        for r_idx, row in enumerate(dataframe_to_rows(all_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_all.cell(row=r_idx+1, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        ws_all.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_all.column_dimensions[col].width = 15

        # Sheet 5: Part Summary
        ws_part_summary = wb.create_sheet('Part_Summary', 4)
        ws_part_summary['A1'] = 'Forecast Accuracy by Part Number'
        ws_part_summary['A1'].font = Font(bold=True, size=14)

        part_summary = comparison_df.groupby('Part_number').agg({
            'Forecast': 'sum',
            'Actual': 'sum',
            'Error': 'sum',
            'Percentage_Error': 'mean',
            'Year_Q': 'count'
        }).reset_index()
        part_summary.columns = ['Part_number', 'Total_Forecast', 'Total_Actual', 'Total_Error', 'Avg_Error_%', 'Quarters']
        part_summary = part_summary.sort_values('Avg_Error_%', key=abs, ascending=False)

        for r_idx, row in enumerate(dataframe_to_rows(part_summary, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_part_summary.cell(row=r_idx+1, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        ws_part_summary.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_part_summary.column_dimensions[col].width = 18

        # Sheet 6: Recommendations
        ws_recommendations = wb.create_sheet('Recommendations', 5)
        ws_recommendations['A1'] = 'RECOMMENDED ACTIONS'
        ws_recommendations['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws_recommendations['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        avg_error = comparison_df['Percentage_Error'].mean()
        extreme_errors = (comparison_df['Percentage_Error'].abs() > 100).sum()

        recommendations = [
            ['', ''],
            ['DIAGNOSIS', ''],
        ]

        if over_count > under_count * 1.2:
            recommendations.extend([
                ['Bias Type:', 'SYSTEMATIC OVER-FORECASTING'],
                ['Severity:', 'Models consistently predict HIGHER than actual demand'],
                ['', ''],
                ['ROOT CAUSES', ''],
                ['1.', 'Models are too optimistic'],
                ['2.', 'Using outdated "good years" as baseline'],
                ['3.', 'Growth rate assumptions too aggressive'],
                ['', ''],
                ['CORRECTION FACTOR', ''],
                ['Suggested multiplier:', f"{1/(1 + comparison_df['Percentage_Error'].median()/100):.3f}x"],
                ['', ''],
            ])
        elif under_count > over_count * 1.2:
            recommendations.extend([
                ['Bias Type:', 'SYSTEMATIC UNDER-FORECASTING'],
                ['Severity:', 'Models consistently predict LOWER than actual demand'],
                ['', ''],
                ['ROOT CAUSES', ''],
                ['1.', 'Models are too pessimistic'],
                ['2.', 'Missing growth signals in data'],
                ['3.', 'Underestimating market trends'],
                ['', ''],
                ['CORRECTION FACTOR', ''],
                ['Suggested multiplier:', f"{1/(1 + comparison_df['Percentage_Error'].median()/100):.3f}x"],
                ['', ''],
            ])
        else:
            recommendations.extend([
                ['Bias Type:', 'BALANCED (no systematic bias)'],
                ['Status:', '✅ Forecasts are balanced between over/under'],
                ['', ''],
            ])

        recommendations.extend([
            ['ACCURACY ASSESSMENT', ''],
        ])

        if abs(avg_error) > 50:
            recommendations.append(['Overall:', f'❌ CRITICAL - {abs(avg_error):.1f}% average error'])
            recommendations.append(['Action:', 'Models fundamentally misaligned with demand'])
        elif abs(avg_error) > 25:
            recommendations.append(['Overall:', f'⚠️ WARNING - {abs(avg_error):.1f}% average error'])
            recommendations.append(['Action:', 'Significant alignment issues need attention'])
        else:
            recommendations.append(['Overall:', f'✅ ACCEPTABLE - {abs(avg_error):.1f}% average error'])

        recommendations.extend([
            ['', ''],
            ['EXTREME OUTLIERS', ''],
            ['Count:', f"{extreme_errors:,} parts with >100% error ({extreme_errors/len(comparison_df)*100:.1f}%)"],
        ])

        if extreme_errors / len(comparison_df) > 0.05:
            recommendations.extend([
                ['Status:', '❌ Too many extreme misses'],
                ['Action 1:', 'Segment products by behavior (stable vs volatile)'],
                ['Action 2:', 'Use different models for different product types'],
                ['Action 3:', 'Review high-value parts manually'],
            ])

        recommendations.extend([
            ['', ''],
            ['IMMEDIATE ACTIONS', ''],
            ['1.', 'Apply bias correction file to future forecasts'],
            ['2.', 'Review Top_Errors sheet - investigate top 20 parts'],
            ['3.', 'Flag items with >50% error for manual review'],
            ['4.', 'Segment parts by ABC classification'],
            ['5.', 'Implement safety stock for high-error parts'],
            ['', ''],
            ['LONG-TERM IMPROVEMENTS', ''],
            ['1.', 'Update monthly with new actuals'],
            ['2.', 'Track accuracy trends over time'],
            ['3.', 'Adjust models for different part behaviors'],
            ['4.', 'Capture leading indicators (not just history)'],
            ['5.', 'Review forecasting methodology quarterly'],
        ])

        for r_idx, row_data in enumerate(recommendations, 1):
            ws_recommendations[f'A{r_idx}'] = row_data[0]
            ws_recommendations[f'B{r_idx}'] = row_data[1]
            if row_data[0] and row_data[0].isupper() and len(row_data[0]) > 3:
                ws_recommendations[f'A{r_idx}'].font = Font(bold=True, color='FFFFFF')
                ws_recommendations[f'A{r_idx}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        ws_recommendations.column_dimensions['A'].width = 25
        ws_recommendations.column_dimensions['B'].width = 60

        # Save workbook
        wb.save(excel_path)

    except Exception as e:
        print(f"⚠️ Error creating Excel analysis: {e}")
        print("   CSV file was saved successfully")

# Parse command line arguments
parser = argparse.ArgumentParser(description='Compare forecasts with actual sales to create accuracy report')
parser.add_argument('--old-forecast', required=True, help='Path to old forecast CSV file')
parser.add_argument('--actual-sales', required=True, help='Path to actual sales Excel/CSV file')
parser.add_argument('--output', default='output/forecast_accuracy_report.csv', help='Output path for accuracy report')
args = parser.parse_args()

print("="*70)
print("FORECAST COMPARISON TOOL")
print("="*70)
print(f"Loading forecast file: {args.old_forecast}")
print(f"Loading actual sales file: {args.actual_sales}")
print()

# Load the files
try:
    forecast_data = pd.read_csv(args.old_forecast)
    print(f"✅ Loaded forecast data: {len(forecast_data)} parts")
except Exception as e:
    print(f"❌ Error loading forecast file: {e}")
    sys.exit(1)

try:
    if args.actual_sales.endswith('.xlsx'):
        actual_data = pd.read_excel(args.actual_sales)
    else:
        actual_data = pd.read_csv(args.actual_sales)
    print(f"✅ Loaded actual sales data: {len(actual_data)} rows")
except Exception as e:
    print(f"❌ Error loading actual sales file: {e}")
    sys.exit(1)

print("\n" + "="*70)
print("PROCESSING DATA")
print("="*70)

# Convert date to datetime and extract quarter and year
actual_data['Date'] = pd.to_datetime(actual_data['Date'])
actual_data['Quarter'] = actual_data['Date'].dt.quarter
actual_data['Year'] = actual_data['Date'].dt.year

# Aggregate actual sales by part number, quarter, and year
actual_sales = actual_data.groupby(['Part_number', 'Year', 'Quarter']).agg({'Value': 'sum'}).reset_index()
actual_sales['Quarter'] = actual_sales['Quarter'].apply(lambda x: 'Q' + str(x))
actual_sales['Year_Q'] = actual_sales['Year'].astype(str).str[-2:] + actual_sales['Quarter']
actual_sales.drop(['Year', 'Quarter'], axis=1, inplace=True)
print(f"✅ Aggregated actual sales by quarter: {len(actual_sales)} part-quarter combinations")

# Detect forecasted columns (any column matching pattern like 25Q1, 26Q2, etc.)
forecasted_columns = [col for col in forecast_data.columns if col[0:2].isdigit() and col[2] == 'Q']
print(f"✅ Detected forecast columns: {forecasted_columns}")

if not forecasted_columns:
    print("❌ No forecast columns found (expected format: 25Q1, 26Q2, etc.)")
    sys.exit(1)

# Prepare comparison dataframe
forecast_long = forecast_data[['Part_number'] + forecasted_columns].melt(
    id_vars='Part_number', 
    value_name='Forecasted_Value', 
    var_name='Year_Q'
)

# Merge actual sales and forecasted values
comparison_df = pd.merge(forecast_long, actual_sales, on=['Part_number', 'Year_Q'], how='inner')
print(f"✅ Matched {len(comparison_df)} forecast-actual pairs")

if len(comparison_df) == 0:
    print("\n❌ No matching data found!")
    print("   Possible issues:")
    print("   - Part numbers don't match between files")
    print("   - Quarter formats don't match (check Year_Q column)")
    print("   - Date ranges don't overlap")
    sys.exit(1)

# Calculate accuracy metrics
comparison_df['Error'] = comparison_df['Forecasted_Value'] - comparison_df['Value']
comparison_df['Absolute_Error'] = comparison_df['Error'].abs()
comparison_df['Percentage_Error'] = (comparison_df['Error'] / comparison_df['Value']) * 100

# Rename columns to match expected format
comparison_df.rename(columns={
    'Forecasted_Value': 'Forecast',
    'Value': 'Actual'
}, inplace=True)

# Print results
print("\n" + "="*70)
print("FORECAST ACCURACY ANALYSIS")
print("="*70)
print(f"\nTotal comparisons: {len(comparison_df)}")
print(f"\nOverall Statistics:")
print(f"  Average error: {comparison_df['Percentage_Error'].mean():.2f}%")
print(f"  Median error: {comparison_df['Percentage_Error'].median():.2f}%")
print(f"  Max over-forecast: {comparison_df['Percentage_Error'].max():.2f}%")
print(f"  Max under-forecast: {comparison_df['Percentage_Error'].min():.2f}%")

print(f"\nAccuracy by Quarter:")
quarter_accuracy = comparison_df.groupby('Year_Q')['Percentage_Error'].agg(['mean', 'median', 'std'])
print(quarter_accuracy)

print(f"\nTop 10 Biggest Forecast Errors:")
top_errors = comparison_df.nlargest(10, 'Absolute_Error')[['Part_number', 'Year_Q', 'Forecast', 'Actual', 'Percentage_Error']]
print(top_errors)

# Save results with correct column names
output_df = comparison_df[['Part_number', 'Year_Q', 'Forecast', 'Actual', 'Error', 'Percentage_Error']]
output_df.to_csv(args.output, index=False)
print(f"\n✅ Full report saved to: {args.output}")

# Calculate bias metrics before Excel creation
over_count = (comparison_df['Error'] > 0).sum()
under_count = (comparison_df['Error'] < 0).sum()
extreme_errors = (comparison_df['Percentage_Error'].abs() > 100).sum()

# Create detailed Excel analysis report
if EXCEL_AVAILABLE:
    excel_output = args.output.replace('.csv', '_analysis.xlsx')
    create_accuracy_analysis_excel(
        comparison_df, 
        quarter_accuracy,
        excel_output,
        args.old_forecast,
        args.actual_sales,
        over_count,
        under_count
    )
    print(f"✅ Detailed analysis report saved to: {excel_output}")
else:
    print("⚠️ Skipping Excel analysis report (openpyxl not installed)")

# Additional analysis: Model-demand alignment diagnosis
print("\n" + "="*70)
print("MODEL-DEMAND ALIGNMENT DIAGNOSIS")
print("="*70)

print(f"\nBias Detection:")
print(f"  Over-forecasted: {over_count:,} ({over_count/len(comparison_df)*100:.1f}%)")
print(f"  Under-forecasted: {under_count:,} ({under_count/len(comparison_df)*100:.1f}%)")

if over_count > under_count * 1.2:
    print("\n⚠️ SYSTEMATIC OVER-FORECASTING DETECTED")
    print("   Your models consistently predict HIGHER than actual demand")
    print("\n   Root Cause: Model-Demand Misalignment")
    print("   → Models are TOO OPTIMISTIC")
    print("   → Check if using outdated 'good years' as baseline")
    print("   → Reduce growth rate assumptions")
    print(f"\n   Suggested correction factor: {1/(1 + comparison_df['Percentage_Error'].median()/100):.3f}x")
elif under_count > over_count * 1.2:
    print("\n⚠️ SYSTEMATIC UNDER-FORECASTING DETECTED")
    print("   Your models consistently predict LOWER than actual demand")
    print("\n   Root Cause: Model-Demand Misalignment")
    print("   → Models are TOO PESSIMISTIC")
    print("   → Missing growth signals in data")
    print("   → May be underestimating market trends")
    print(f"\n   Suggested correction factor: {1/(1 + comparison_df['Percentage_Error'].median()/100):.3f}x")
else:
    print("\n✅ No systematic bias - forecasts are balanced")

avg_error = comparison_df['Percentage_Error'].mean()
print(f"\nAccuracy Assessment:")
if abs(avg_error) > 50:
    print(f"   ❌ CRITICAL: {abs(avg_error):.1f}% average error")
    print("      Models fundamentally misaligned with demand")
elif abs(avg_error) > 25:
    print(f"   ⚠️ WARNING: {abs(avg_error):.1f}% average error")
    print("      Significant alignment issues")
else:
    print(f"   ✅ ACCEPTABLE: {abs(avg_error):.1f}% average error")

print(f"\nExtreme Outliers: {extreme_errors:,} ({extreme_errors/len(comparison_df)*100:.1f}% with >100% error)")
if extreme_errors / len(comparison_df) > 0.05:
    print("   ❌ Too many extreme misses - missing key demand signals")
    print("   → Segment products by behavior (stable vs volatile)")
    print("   → Use different models for different product types")

print("\n" + "="*70)
print("RECOMMENDED ACTIONS TO FIX ALIGNMENT")
print("="*70)
print("""
The phrase "Fix the model-demand alignment" means:

1. **Input Data Quality**
   - Are models using LEADING indicators (future signals)?
   - Or LAGGING indicators (past performance only)?
   - Check: Seasonality, promotions, market events

2. **Model Assumptions**
   - Trained on representative historical periods?
   - Account for market changes (trends, competition)?
   - Right forecasting method for each product type?

3. **Demand Signal Capture**
   - Fast-moving items: Short-term reactive models
   - Slow-moving items: Intermittent demand models
   - New products: Can't rely on history alone

4. **Quick Fixes**
   ✓ Flag items with >50% error for manual review
   ✓ Apply bias correction (adjust systematically high/low)
   ✓ Segment by ABC: A-items need sophisticated models
   ✓ Review forecasting methodology for high-error segments
   ✓ Implement safety stock instead of "perfect" forecasts

Goal: ALIGNMENT between what models predict and what customers actually buy.
Not perfection - just better alignment with reality.
""")
