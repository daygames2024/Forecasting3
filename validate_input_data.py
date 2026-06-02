"""
Input Data Validation Tool
Validates and formats sales data and maintenance events before forecasting

Usage:
    python validate_input_data.py --sales "sales_data.xlsx"
    python validate_input_data.py --sales "sales_data.xlsx" --events "maintenance_events.csv"
    python validate_input_data.py --sales "sales_data.xlsx" --fix
"""

import pandas as pd
import argparse
import sys
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def print_header(text):
    """Print formatted section header"""
    print("\n" + "="*70)
    print(text)
    print("="*70)

def print_success(text):
    """Print success message"""
    print(f"[OK] {text}")

def print_warning(text):
    """Print warning message"""
    print(f"[WARNING] {text}")

def print_error(text):
    """Print error message"""
    print(f"[ERROR] {text}")

def validate_sales_data(filepath, fix=False):
    """
    Validate sales data file format

    Args:
        filepath: Path to sales data file
        fix: If True, attempt to fix issues and save corrected file

    Returns:
        tuple: (is_valid, dataframe, issues_list, output_filepath)
    """
    print_header("VALIDATING SALES DATA")
    print(f"File: {filepath}\n")

    issues = []
    is_valid = True
    output_filepath = filepath  # Track the output file path

    # Check if file exists
    if not os.path.exists(filepath):
        print_error(f"File not found: {filepath}")
        return False, None, [f"File not found: {filepath}"], filepath

    # Load file
    try:
        if filepath.endswith('.xlsx'):
            df = pd.read_excel(filepath)
            print_success(f"Loaded Excel file: {len(df)} rows")
        elif filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
            print_success(f"Loaded CSV file: {len(df)} rows")
        else:
            print_error("File must be .xlsx or .csv")
            return False, None, ["Unsupported file format"], filepath
    except Exception as e:
        print_error(f"Error loading file: {e}")
        return False, None, [f"Error loading file: {e}"], filepath

    # Check required columns
    required_columns = ['Date', 'Part_number', 'Value']
    print("\nChecking required columns...")

    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        is_valid = False
        issue = f"Missing required columns: {missing_columns}"
        issues.append(issue)
        print_error(issue)
        print(f"   Found columns: {list(df.columns)}")
        print(f"   Expected columns: {required_columns}")
    else:
        print_success(f"All required columns found: {required_columns}")

    if not is_valid and not fix:
        return is_valid, df, issues, output_filepath

    # Validate Date column
    print("\nValidating Date column...")
    original_dates = df['Date'].copy()

    try:
        # Attempt to parse dates
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

        # Check for unparseable dates
        null_dates = df['Date'].isna()
        if null_dates.any():
            num_null = null_dates.sum()
            issue = f"{num_null} dates could not be parsed"
            issues.append(issue)
            print_warning(issue)

            # Show examples of bad dates
            bad_date_examples = original_dates[null_dates].head(5).tolist()
            print(f"   Examples of unparseable dates: {bad_date_examples}")
            is_valid = False
        else:
            print_success("All dates parsed successfully")

        # Show date range
        if not df['Date'].isna().all():
            min_date = df['Date'].min()
            max_date = df['Date'].max()
            print(f"   Date range: {min_date.date()} to {max_date.date()}")

            # Show quarter distribution
            df['Quarter'] = df['Date'].dt.quarter
            df['Year'] = df['Date'].dt.year
            df['Year_Q'] = df['Year'].astype(str).str[-2:] + 'Q' + df['Quarter'].astype(str)

            print("\n   Quarter distribution:")
            quarter_counts = df['Year_Q'].value_counts().sort_index()
            for quarter, count in quarter_counts.items():
                print(f"      {quarter}: {count:,} rows")

            df.drop(['Quarter', 'Year', 'Year_Q'], axis=1, inplace=True)

    except Exception as e:
        issue = f"Error parsing dates: {e}"
        issues.append(issue)
        print_error(issue)
        is_valid = False

    # Validate Part_number column
    print("\nValidating Part_number column...")

    null_parts = df['Part_number'].isna()
    if null_parts.any():
        num_null = null_parts.sum()
        issue = f"{num_null} rows have missing Part_number"
        issues.append(issue)
        print_warning(issue)
        is_valid = False
    else:
        print_success("No missing Part_numbers")

    unique_parts = df['Part_number'].nunique()
    print(f"   Unique parts: {unique_parts:,}")

    # Show top 5 parts by transaction count
    top_parts = df['Part_number'].value_counts().head(5)
    print("\n   Top 5 parts by transaction count:")
    for part, count in top_parts.items():
        print(f"      {part}: {count:,} transactions")

    # Validate Value column
    print("\nValidating Value column...")

    # Check for non-numeric values
    try:
        df['Value'] = pd.to_numeric(df['Value'], errors='coerce')

        null_values = df['Value'].isna()
        if null_values.any():
            num_null = null_values.sum()
            issue = f"{num_null} rows have non-numeric or missing Values"
            issues.append(issue)
            print_warning(issue)
            is_valid = False
        else:
            print_success("All Values are numeric")

        # Check for negative values
        negative_values = df['Value'] < 0
        if negative_values.any():
            num_negative = negative_values.sum()
            issue = f"{num_negative} rows have negative Values"
            issues.append(issue)
            print_warning(issue)

        # Show value statistics
        if not df['Value'].isna().all():
            print(f"   Min value: {df['Value'].min()}")
            print(f"   Max value: {df['Value'].max()}")
            print(f"   Mean value: {df['Value'].mean():.2f}")
            print(f"   Median value: {df['Value'].median():.2f}")

    except Exception as e:
        issue = f"Error validating Values: {e}"
        issues.append(issue)
        print_error(issue)
        is_valid = False

    # Check for duplicates
    print("\nChecking for duplicates...")
    duplicates = df.duplicated(subset=['Date', 'Part_number'], keep=False)
    if duplicates.any():
        num_duplicates = duplicates.sum()
        issue = f"{num_duplicates} duplicate rows (same Date + Part_number)"
        issues.append(issue)
        print_warning(issue)
        print("   Note: Duplicates are OK if they represent multiple transactions per day")
    else:
        print_success("No duplicate Date + Part_number combinations")

    # Summary
    print_header("VALIDATION SUMMARY")
    if is_valid:
        print_success("Sales data is valid and ready for forecasting!")
    else:
        print_error(f"Found {len(issues)} issue(s) that need attention")
        for i, issue in enumerate(issues, 1):
            print(f"   {i}. {issue}")

    # Fix and save if requested
    if fix and len(issues) > 0:
        print_header("FIXING ISSUES")

        # Remove rows with null dates, parts, or values
        original_len = len(df)
        df = df.dropna(subset=['Date', 'Part_number', 'Value'])
        removed = original_len - len(df)

        if removed > 0:
            print_warning(f"Removed {removed} rows with missing data")

        # Standardize date format
        df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        print_success("Standardized dates to YYYY-MM-DD format")

        # Sort by date and part
        df = df.sort_values(['Part_number', 'Date'])
        print_success("Sorted data by Part_number and Date")

        # Save fixed file
        output_path = filepath.replace('.xlsx', '_FIXED.xlsx').replace('.csv', '_FIXED.csv')
        output_filepath = output_path  # Update the output filepath

        try:
            if output_path.endswith('.xlsx'):
                # Save to Excel with explicit writer to ensure proper file creation
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sales_Data')
                # File is automatically closed and flushed when exiting context manager
            else:
                df.to_csv(output_path, index=False)

            print_success(f"Fixed file saved to: {output_path}")
            print("\n   Next steps:")
            print(f"   1. Review the fixed file: {output_path}")
            print(f"   2. Run forecast with: python -m src.app --input \"{output_path}\"")

        except Exception as e:
            print_error(f"Error saving fixed file: {e}")

    return is_valid, df, issues, output_filepath


def validate_events_data(filepath, fix=False):
    """
    Validate maintenance events file format

    Args:
        filepath: Path to events CSV file
        fix: If True, attempt to fix issues and save corrected file

    Returns:
        tuple: (is_valid, dataframe, issues_list)
    """
    print_header("VALIDATING MAINTENANCE EVENTS")
    print(f"File: {filepath}\n")

    issues = []
    is_valid = True

    # Check if file exists
    if not os.path.exists(filepath):
        print_error(f"File not found: {filepath}")
        return False, None, [f"File not found: {filepath}"]

    # Load file
    try:
        df = pd.read_csv(filepath)
        print_success(f"Loaded events file: {len(df)} events")
    except Exception as e:
        print_error(f"Error loading file: {e}")
        return False, None, [f"Error loading file: {e}"]

    # Check required columns
    required_columns = ['Check', 'Start Date', 'End Date']
    print("\nChecking required columns...")

    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        is_valid = False
        issue = f"Missing required columns: {missing_columns}"
        issues.append(issue)
        print_error(issue)
        print(f"   Found columns: {list(df.columns)}")
        print(f"   Expected columns: {required_columns}")
    else:
        print_success(f"All required columns found: {required_columns}")

    if not is_valid and not fix:
        return is_valid, df, issues

    # Validate Check column
    print("\nValidating Check (event name) column...")

    null_checks = df['Check'].isna()
    if null_checks.any():
        num_null = null_checks.sum()
        issue = f"{num_null} events have missing Check names"
        issues.append(issue)
        print_warning(issue)
        is_valid = False
    else:
        print_success("No missing event names")

    unique_checks = df['Check'].value_counts()
    print(f"\n   Event types found:")
    for check, count in unique_checks.items():
        print(f"      {check}: {count} events")

    # Validate dates
    print("\nValidating Start Date and End Date...")

    original_start = df['Start Date'].copy()
    original_end = df['End Date'].copy()

    try:
        # Parse start dates
        df['Start Date'] = pd.to_datetime(df['Start Date'], errors='coerce')
        null_start = df['Start Date'].isna()
        if null_start.any():
            num_null = null_start.sum()
            issue = f"{num_null} Start Dates could not be parsed"
            issues.append(issue)
            print_warning(issue)

            bad_examples = original_start[null_start].head(5).tolist()
            print(f"   Examples: {bad_examples}")
            is_valid = False
        else:
            print_success("All Start Dates parsed successfully")

        # Parse end dates
        df['End Date'] = pd.to_datetime(df['End Date'], errors='coerce')
        null_end = df['End Date'].isna()
        if null_end.any():
            num_null = null_end.sum()
            issue = f"{num_null} End Dates could not be parsed"
            issues.append(issue)
            print_warning(issue)

            bad_examples = original_end[null_end].head(5).tolist()
            print(f"   Examples: {bad_examples}")
            is_valid = False
        else:
            print_success("All End Dates parsed successfully")

        # Check date logic (end after start)
        if not df['Start Date'].isna().any() and not df['End Date'].isna().any():
            invalid_ranges = df['End Date'] < df['Start Date']
            if invalid_ranges.any():
                num_invalid = invalid_ranges.sum()
                issue = f"{num_invalid} events have End Date before Start Date"
                issues.append(issue)
                print_warning(issue)

                print("\n   Invalid events:")
                for idx in df[invalid_ranges].index:
                    row = df.loc[idx]
                    print(f"      {row['Check']}: {row['Start Date'].date()} to {row['End Date'].date()}")
                is_valid = False
            else:
                print_success("All events have valid date ranges (End after Start)")

        # Show date range
        if not df['Start Date'].isna().all():
            min_date = df['Start Date'].min()
            max_date = df['End Date'].max()
            print(f"\n   Events date range: {min_date.date()} to {max_date.date()}")

    except Exception as e:
        issue = f"Error parsing dates: {e}"
        issues.append(issue)
        print_error(issue)
        is_valid = False

    # Summary
    print_header("VALIDATION SUMMARY")
    if is_valid:
        print_success("Maintenance events data is valid!")
    else:
        print_error(f"Found {len(issues)} issue(s) that need attention")
        for i, issue in enumerate(issues, 1):
            print(f"   {i}. {issue}")

    # Fix and save if requested
    if fix and len(issues) > 0:
        print_header("FIXING ISSUES")

        # Remove rows with null data
        original_len = len(df)
        df = df.dropna(subset=['Check', 'Start Date', 'End Date'])
        removed = original_len - len(df)

        if removed > 0:
            print_warning(f"Removed {removed} events with missing data")

        # Remove invalid date ranges
        valid_ranges = df['End Date'] >= df['Start Date']
        df = df[valid_ranges]
        removed_invalid = (~valid_ranges).sum()

        if removed_invalid > 0:
            print_warning(f"Removed {removed_invalid} events with invalid date ranges")

        # Standardize date format
        df['Start Date'] = df['Start Date'].dt.strftime('%d-%m-%y')
        df['End Date'] = df['End Date'].dt.strftime('%d-%m-%y')
        print_success("Standardized dates to DD-MM-YY format")

        # Sort by start date
        df = df.sort_values('Start Date')
        print_success("Sorted events by Start Date")

        # Save fixed file
        output_path = filepath.replace('.csv', '_FIXED.csv')

        try:
            df.to_csv(output_path, index=False)
            print_success(f"Fixed file saved to: {output_path}")
            print("\n   Next steps:")
            print(f"   1. Review the fixed file: {output_path}")
            print(f"   2. Run forecast with: python -m src.app --input \"sales_data.xlsx\" --events-file \"{output_path}\"")

        except Exception as e:
            print_error(f"Error saving fixed file: {e}")

    return is_valid, df, issues


def compare_date_ranges(sales_df, events_df):
    """Compare date ranges between sales and events data"""
    print_header("COMPARING SALES AND EVENTS DATE RANGES")

    if sales_df is None or events_df is None:
        print_warning("Cannot compare - one or both files have validation errors")
        return

    try:
        # Get sales date range
        sales_min = sales_df['Date'].min()
        sales_max = sales_df['Date'].max()

        # Get events date range
        events_min = events_df['Start Date'].min()
        events_max = events_df['End Date'].max()

        print(f"Sales data range:  {sales_min.date()} to {sales_max.date()}")
        print(f"Events data range: {events_min.date()} to {events_max.date()}")

        # Check overlap
        if events_max < sales_min:
            print_warning("Events end BEFORE sales data starts - no overlap!")
        elif events_min > sales_max:
            print_warning("Events start AFTER sales data ends - no overlap!")
        else:
            print_success("Date ranges overlap - events will be applied to forecasts")

            # Check if events cover forecast period
            # Typically forecast 6-8 quarters ahead from last sales date
            forecast_start = sales_max
            forecast_end = sales_max + pd.DateOffset(months=18)  # 6 quarters

            print(f"\nExpected forecast range: {forecast_start.date()} to {forecast_end.date()}")

            if events_max < forecast_start:
                print_warning("Events don't cover forecast period - forecasts will have no events applied")
            elif events_min > forecast_end:
                print_warning("Events start after forecast period - forecasts will have no events applied")
            else:
                print_success("Events cover at least part of forecast period")

    except Exception as e:
        print_error(f"Error comparing dates: {e}")


def create_validation_summary_excel(sales_filepath, sales_df, sales_issues, events_df=None, events_issues=None):
    """
    Create Excel summary of validation results and add as new sheet to sales workbook

    Args:
        sales_filepath: Path to original sales file
        sales_df: Validated sales dataframe
        sales_issues: List of sales validation issues
        events_df: Validated events dataframe (optional)
        events_issues: List of events validation issues (optional)
    """
    print_header("CREATING VALIDATION SUMMARY SHEET")

    if not sales_filepath.endswith('.xlsx'):
        print_warning("Summary sheet only works with .xlsx files, not .csv")
        return

    try:
        # Create summary data
        summary_data = []

        # Header
        summary_data.append(['FORECAST INPUT DATA VALIDATION SUMMARY', ''])
        summary_data.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_data.append(['', ''])

        # Sales Data Summary
        summary_data.append(['SALES DATA VALIDATION', ''])
        summary_data.append(['File:', sales_filepath])
        summary_data.append(['Total Rows:', len(sales_df)])
        summary_data.append(['', ''])

        # Date Analysis
        date_col = sales_df['Date']

        # If dates are strings (after --fix), convert back to datetime for analysis
        if date_col.dtype == 'object':
            try:
                date_col = pd.to_datetime(date_col, errors='coerce')
            except:
                date_col = None

        if date_col is not None and (date_col.dtype == 'datetime64[ns]' or isinstance(date_col.iloc[0] if len(date_col) > 0 else None, pd.Timestamp)):
            summary_data.append(['Date Range:', ''])
            summary_data.append(['  Min Date:', date_col.min().strftime('%Y-%m-%d')])
            summary_data.append(['  Max Date:', date_col.max().strftime('%Y-%m-%d')])
            summary_data.append(['', ''])

            # Quarter distribution
            temp_df = pd.DataFrame({'Date': date_col})
            temp_df['Quarter'] = temp_df['Date'].dt.quarter
            temp_df['Year'] = temp_df['Date'].dt.year
            temp_df['Year_Q'] = temp_df['Year'].astype(str).str[-2:] + 'Q' + temp_df['Quarter'].astype(str)

            quarter_counts = temp_df['Year_Q'].value_counts().sort_index()

            summary_data.append(['Quarter Distribution:', 'Row Count'])
            for quarter, count in quarter_counts.items():
                summary_data.append([f'  {quarter}', count])
            summary_data.append(['', ''])

        # Part Analysis
        if 'Part_number' in sales_df.columns:
            summary_data.append(['Part Numbers:', ''])
            summary_data.append(['  Unique Parts:', sales_df['Part_number'].nunique()])
            summary_data.append(['', ''])

            top_parts = sales_df['Part_number'].value_counts().head(10)
            summary_data.append(['Top 10 Parts by Transaction Count:', 'Transactions'])
            for part, count in top_parts.items():
                summary_data.append([f'  {part}', count])
            summary_data.append(['', ''])

        # Value Statistics
        if 'Value' in sales_df.columns:
            summary_data.append(['Value Statistics:', ''])
            summary_data.append(['  Min:', sales_df['Value'].min()])
            summary_data.append(['  Max:', sales_df['Value'].max()])
            summary_data.append(['  Mean:', round(sales_df['Value'].mean(), 2)])
            summary_data.append(['  Median:', sales_df['Value'].median()])
            summary_data.append(['  Total:', sales_df['Value'].sum()])
            summary_data.append(['', ''])

        # Validation Issues
        summary_data.append(['Validation Issues:', ''])
        if len(sales_issues) == 0:
            summary_data.append(['  ✅ No issues found', ''])
        else:
            for i, issue in enumerate(sales_issues, 1):
                summary_data.append([f'  ⚠️  Issue {i}:', issue])
        summary_data.append(['', ''])

        # Events Summary (if provided)
        if events_df is not None:
            summary_data.append(['MAINTENANCE EVENTS VALIDATION', ''])
            summary_data.append(['Total Events:', len(events_df)])
            summary_data.append(['', ''])

            if 'Check' in events_df.columns:
                event_types = events_df['Check'].value_counts()
                summary_data.append(['Event Types:', 'Count'])
                for event, count in event_types.items():
                    summary_data.append([f'  {event}', count])
                summary_data.append(['', ''])

            if 'Start Date' in events_df.columns and events_df['Start Date'].dtype == 'datetime64[ns]':
                summary_data.append(['Event Date Range:', ''])
                summary_data.append(['  Earliest Start:', events_df['Start Date'].min().strftime('%Y-%m-%d')])
                summary_data.append(['  Latest End:', events_df['End Date'].max().strftime('%Y-%m-%d')])
                summary_data.append(['', ''])

            if events_issues and len(events_issues) > 0:
                summary_data.append(['Events Validation Issues:', ''])
                for i, issue in enumerate(events_issues, 1):
                    summary_data.append([f'  ⚠️  Issue {i}:', issue])
                summary_data.append(['', ''])

        # Forecast Readiness
        summary_data.append(['FORECAST READINESS', ''])
        if len(sales_issues) == 0 and (events_issues is None or len(events_issues) == 0):
            summary_data.append(['Status:', '✅ READY FOR FORECASTING'])
        else:
            summary_data.append(['Status:', '⚠️  HAS ISSUES - Review above'])
        summary_data.append(['', ''])

        # Next Steps
        summary_data.append(['NEXT STEPS', ''])
        summary_data.append(['1.', 'Review validation issues above'])
        summary_data.append(['2.', 'Fix any critical issues'])
        summary_data.append(['3.', 'Run forecast command:'])
        if events_df is not None:
            summary_data.append(['', 'python -m src.app --input "sales_data.xlsx" --events-file "maintenance_events.csv"'])
        else:
            summary_data.append(['', 'python -m src.app --input "sales_data.xlsx"'])

        # Load existing workbook
        wb = load_workbook(sales_filepath)

        # Remove old summary sheet if exists
        if 'Validation_Summary' in wb.sheetnames:
            del wb['Validation_Summary']

        # Create new sheet at the beginning
        ws = wb.create_sheet('Validation_Summary', 0)

        print(f"   Writing {len(summary_data)} rows to Excel...")

        # Write data directly from list (no DataFrame conversion needed)
        rows_written = 0
        for r_idx, row_data in enumerate(summary_data, 1):
            # row_data is a list with 2 elements: [metric, value]
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                rows_written += 1

                # Formatting
                # Header row (first row)
                if r_idx == 1:
                    cell.font = Font(bold=True, size=14, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left')

                # Section headers (all caps in column A)
                elif value and isinstance(value, str) and value.isupper() and c_idx == 1:
                    cell.font = Font(bold=True, size=12, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

                # Sub-headers (ends with ':')
                elif value and isinstance(value, str) and value.endswith(':') and not value.startswith(' '):
                    cell.font = Font(bold=True)

                # Status indicators
                elif value and isinstance(value, str):
                    if '✅' in value:
                        cell.font = Font(color='006100')  # Green
                    elif '⚠️' in value or 'WARNING' in value:
                        cell.font = Font(color='C65911')  # Orange
                    elif '❌' in value or 'FAILED' in value:
                        cell.font = Font(color='C00000')  # Red

        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30

        print(f"   Cells written: {rows_written}")

        # Save workbook
        wb.save(sales_filepath)

        print_success(f"Validation summary added to: {sales_filepath}")
        print(f"   Sheet name: 'Validation_Summary' (first tab)")
        print("\n   Open the Excel file to view the summary")
        print("   This summary will be included when you run forecasts with this file")

    except ImportError as e:
        print_error(f"Missing required library: {e}")
        print("   Install with: python -m pip install openpyxl")
        print("   Continuing without summary sheet...")
    except Exception as e:
        print_error(f"Could not create summary sheet: {e}")
        print(f"   Error type: {type(e).__name__}")
        import traceback
        print(f"   Details: {traceback.format_exc()}")
        print("   Continuing without summary sheet...")


def main():
    parser = argparse.ArgumentParser(
        description='Validate and fix input data for forecasting',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Validate sales data only
    python validate_input_data.py --sales "sales_data.xlsx"

    # Validate both sales and events
    python validate_input_data.py --sales "sales_data.xlsx" --events "maintenance_events.csv"

    # Validate and automatically fix issues
    python validate_input_data.py --sales "sales_data.xlsx" --fix

    # Validate, fix, and compare date ranges
    python validate_input_data.py --sales "sales_data.xlsx" --events "maintenance_events.csv" --fix
        """
    )

    parser.add_argument('--sales', required=True, help='Path to sales data file (Excel or CSV)')
    parser.add_argument('--events', help='Path to maintenance events CSV file (optional)')
    parser.add_argument('--fix', action='store_true', help='Attempt to fix issues and save corrected files')

    args = parser.parse_args()

    print_header("INPUT DATA VALIDATION TOOL")
    print(f"Sales file: {args.sales}")
    if args.events:
        print(f"Events file: {args.events}")
    if args.fix:
        print("Mode: FIX (will attempt to correct issues)")
    else:
        print("Mode: VALIDATE ONLY (use --fix to automatically correct issues)")

    # Validate sales data
    sales_valid, sales_df, sales_issues, sales_output_path = validate_sales_data(args.sales, fix=args.fix)

    # Validate events data if provided
    events_valid = True
    events_df = None
    events_issues = None  # Initialize events_issues
    if args.events:
        events_valid, events_df, events_issues = validate_events_data(args.events, fix=args.fix)

        # Compare date ranges if both are valid
        if sales_valid and events_valid:
            compare_date_ranges(sales_df, events_df)

    # Final summary
    print_header("FINAL SUMMARY")

    if sales_valid and events_valid:
        print_success("All validations passed!")

        # Create Excel summary sheet - use the output file path (FIXED file if --fix was used)
        if sales_output_path.endswith('.xlsx'):
            create_validation_summary_excel(sales_output_path, sales_df, sales_issues, events_df, events_issues)

        if args.fix:
            print("\n✅ Fixed files are ready for forecasting")
        else:
            print("\n✅ Your data is ready for forecasting")
            print("\nRun forecast with:")
            if args.events:
                print(f"   python -m src.app --input \"{args.sales}\" --events-file \"{args.events}\"")
            else:
                print(f"   python -m src.app --input \"{args.sales}\"")
    else:
        print_error("Validation failed - please fix issues before forecasting")
        if not args.fix:
            print("\nTip: Run with --fix flag to automatically correct common issues:")
            print(f"   python validate_input_data.py --sales \"{args.sales}\" --fix")
        sys.exit(1)


if __name__ == '__main__':
    main()
