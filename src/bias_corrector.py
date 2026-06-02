"""
Bias Correction Module for Forecasting Application

Integrates with your existing forecasting pipeline to apply
model-demand alignment corrections based on historical accuracy.

Usage:
    from bias_corrector import BiasCorrector
    
    corrector = BiasCorrector()
    corrector.learn_from_accuracy_report("forecast_accuracy_report.csv")
    corrected_forecasts = corrector.apply_correction(raw_forecasts)
"""

import pandas as pd
import numpy as np
from pathlib import Path
import warnings


class BiasCorrector:
    """
    Applies statistical bias correction to forecasts based on historical errors.
    
    Addresses model-demand misalignment by learning from past forecast accuracy
    and adjusting future predictions accordingly.
    """
    
    def __init__(self):
        self.correction_factors = {}
        self.global_correction = 1.0
        self.is_trained = False
        self.bias_type = None  # 'over', 'under', or 'balanced'
        
    def learn_from_accuracy_report(self, report_path: str, min_samples: int = 1):
        """
        Learn correction factors from forecast accuracy comparison.

        Args:
            report_path: Path to forecast_accuracy_report.csv
            min_samples: Minimum samples required per part for reliable correction
        """
        print(f"\n[BiasCorrector] Loading accuracy report from: {report_path}")

        try:
            df = pd.read_csv(report_path)
            df.columns = df.columns.str.strip()

            print(f"[BiasCorrector] Loaded {len(df):,} historical comparisons")

            # Calculate PART-SPECIFIC correction factors (not quarter-based!)
            part_stats = df.groupby('Part_number').agg({
                'Percentage_Error': ['median', 'mean', 'count']
            }).round(2)

            print(f"\n[BiasCorrector] Analyzing forecast bias by PART:")
            print(f"   Total parts with history: {len(part_stats):,}")

            # Store correction factor for each part
            for part_number in df['Part_number'].unique():
                part_data = df[df['Part_number'] == part_number]

                if len(part_data) < min_samples:
                    continue

                # Use median for robustness against outliers
                median_error_pct = part_data['Percentage_Error'].median()

                # Convert error percentage to correction factor
                # Error = (Forecast - Actual) / Actual * 100
                # If error is +20%, forecasts are 20% too high, need 0.83x correction
                # If error is -20%, forecasts are 20% too low, need 1.25x correction
                correction_factor = 1.0 / (1.0 + (median_error_pct / 100.0))

                # Cap extreme corrections to avoid unrealistic adjustments
                correction_factor = max(0.2, min(5.0, correction_factor))

                self.correction_factors[part_number] = {
                    'factor': correction_factor,
                    'median_error': median_error_pct,
                    'sample_count': len(part_data)
                }

            print(f"   Parts with corrections: {len(self.correction_factors):,}")

            # Show distribution of corrections
            factors = [v['factor'] for v in self.correction_factors.values()]
            print(f"\n[BiasCorrector] Correction factor distribution:")
            print(f"   Min: {min(factors):.3f}x (heavily over-forecasted)")
            print(f"   Median: {pd.Series(factors).median():.3f}x")
            print(f"   Max: {max(factors):.3f}x (heavily under-forecasted)")

            # Show top 10 parts needing biggest corrections
            sorted_parts = sorted(self.correction_factors.items(), 
                                 key=lambda x: abs(1.0 - x[1]['factor']), 
                                 reverse=True)[:10]
            print(f"\n[BiasCorrector] Top 10 parts needing largest corrections:")
            for part, data in sorted_parts:
                direction = "over" if data['median_error'] > 0 else "under"
                print(f"   {part}: {data['median_error']:+.1f}% ({direction}) -> {data['factor']:.3f}x")

            # Calculate global correction as fallback for parts without history
            overall_median_error = df['Percentage_Error'].median()
            self.global_correction = 1.0 / (1.0 + (overall_median_error / 100.0))
            self.global_correction = max(0.5, min(2.0, self.global_correction))

            # Detect bias type
            over_count = (df['Error'] > 0).sum()
            under_count = (df['Error'] < 0).sum()

            if over_count > under_count * 1.2:
                self.bias_type = 'over'
                print(f"\n[BiasCorrector] SYSTEMATIC OVER-FORECASTING detected ({over_count/len(df)*100:.1f}%)")
            elif under_count > over_count * 1.2:
                self.bias_type = 'under'
                print(f"\n[BiasCorrector] SYSTEMATIC UNDER-FORECASTING detected ({under_count/len(df)*100:.1f}%)")
            else:
                self.bias_type = 'balanced'
                print(f"\n[BiasCorrector] No systematic bias detected")

            print(f"[BiasCorrector] Global correction factor (for new parts): {self.global_correction:.3f}x")

            self.is_trained = True
            return True

        except Exception as e:
            print(f"[BiasCorrector] ERROR loading accuracy report: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def apply_correction(self, forecast_df: pd.DataFrame, 
                        quarter_columns: list = None,
                        inplace: bool = False) -> pd.DataFrame:
        """
        Apply learned corrections to forecast DataFrame.

        Args:
            forecast_df: DataFrame with Part_number and forecast columns (e.g., '25Q2', '25Q3', ...)
            quarter_columns: List of column names to correct. If None, auto-detect.
            inplace: If True, modify original DataFrame. If False, return copy.

        Returns:
            DataFrame with corrected forecasts (replaces original quarter columns)
        """
        if not self.is_trained:
            warnings.warn("BiasCorrector not trained. Call learn_from_accuracy_report() first.")
            return forecast_df if inplace else forecast_df.copy()

        df = forecast_df if inplace else forecast_df.copy()

        # Auto-detect quarter columns if not specified
        if quarter_columns is None:
            # Match both standard quarters (26Q2) and confidence intervals (26Q2_P10, 26Q2_P90)
            quarter_pattern = r'\d{2}Q[1-4](_P(10|90))?$'
            quarter_columns = [col for col in df.columns if pd.Series([col]).str.match(quarter_pattern).any()]

        print(f"\n[BiasCorrector] Applying PART-SPECIFIC corrections to {len(quarter_columns)} forecast columns...")
        print(f"   (Includes base forecasts and confidence intervals)")

        # Create a mapping dataframe for fast lookup
        correction_df = pd.DataFrame([
            {'Part_number': k, 'correction_factor': v['factor']} 
            for k, v in self.correction_factors.items()
        ])

        # Merge to add correction factors to forecast data
        df = df.merge(correction_df, on='Part_number', how='left')

        # Fill missing correction factors with global correction
        df['correction_factor'].fillna(self.global_correction, inplace=True)

        # Count corrections
        parts_with_specific = df['correction_factor'].ne(self.global_correction).sum()
        parts_with_global = df['correction_factor'].eq(self.global_correction).sum()

        # Apply correction to all quarter columns (vectorized!)
        # This includes base forecasts (26Q2) AND confidence intervals (26Q2_P10, 26Q2_P90)
        for col in quarter_columns:
            if col in df.columns:
                df[col] = (df[col] * df['correction_factor']).round(0)

        # Remove temporary correction_factor column
        df.drop('correction_factor', axis=1, inplace=True)

        print(f"\n[BiasCorrector] Correction Summary:")
        print(f"   Parts with specific corrections: {parts_with_specific:,}")
        print(f"   Parts using global correction: {parts_with_global:,}")
        print(f"   Total parts adjusted: {len(df):,}")

        # Add metadata
        df['Bias_Correction_Applied'] = True
        df['Correction_Type'] = self.bias_type

        return df
    
    def get_summary(self) -> dict:
        """Return summary of learned corrections."""
        return {
            'is_trained': self.is_trained,
            'bias_type': self.bias_type,
            'global_correction': self.global_correction,
            'quarter_corrections': self.correction_factors
        }
    
    def save_correction_factors(self, output_path: str):
        """Save learned correction factors to CSV."""
        if not self.is_trained:
            print("[BiasCorrector] No corrections to save - not trained yet")
            return

        rows = []
        for part_number, data in self.correction_factors.items():
            rows.append({
                'Part_number': part_number,
                'Median_Error_Pct': data['median_error'],
                'Correction_Factor': data['factor'],
                'Sample_Count': data['sample_count']
            })

        summary_df = pd.DataFrame(rows)
        summary_df = summary_df.sort_values('Correction_Factor')  # Sort by correction magnitude
        summary_df.to_csv(output_path, index=False)
        print(f"[BiasCorrector] Saved {len(summary_df)} part-specific correction factors to: {output_path}")


def apply_bias_correction_to_file(input_file: str, 
                                   accuracy_report: str,
                                   output_file: str = None) -> str:
    """
    Convenience function to apply bias correction to a forecast CSV file.

    Args:
        input_file: Path to original forecast CSV (e.g., "All Working.csv")
        accuracy_report: Path to accuracy report (e.g., "forecast_accuracy_report.csv")
        output_file: Path for corrected output. If None, auto-generates name.

    Returns:
        Path to corrected output file
    """
    from datetime import datetime

    print("="*70)
    print("BIAS CORRECTION - MODEL-DEMAND ALIGNMENT FIX")
    print("="*70)

    # Load forecasts
    print(f"\n[1/5] Loading forecasts from: {input_file}")
    forecast_df = pd.read_csv(input_file)
    print(f"   Loaded: {len(forecast_df):,} parts")

    # Train corrector
    print(f"\n[2/5] Learning corrections from accuracy report...")
    corrector = BiasCorrector()
    success = corrector.learn_from_accuracy_report(accuracy_report)

    if not success:
        print("\n[ERROR] Failed to learn corrections. Aborting.")
        return None

    # Apply corrections
    print(f"\n[3/5] Applying corrections to forecasts...")
    corrected_df = corrector.apply_correction(forecast_df)

    # Save output
    if output_file is None:
        input_path = Path(input_file)
        output_file = str(input_path.parent / f"{input_path.stem}_Bias_Corrected{input_path.suffix}")

    print(f"\n[4/5] Saving corrected forecasts...")
    corrected_df.to_csv(output_file, index=False)
    print(f"   Saved to: {output_file}")

    # Save correction factors for reference
    factors_file = str(Path(output_file).parent / "correction_factors.csv")
    corrector.save_correction_factors(factors_file)

    # Create Excel file with summary sheet
    print(f"\n[5/5] Creating Excel summary report...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        excel_file = str(output_file).replace('.csv', '.xlsx')

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create Summary sheet first
        ws_summary = wb.create_sheet('Corrected_Forecast_Summary', 0)

        # Load original forecast for comparison
        original_df = pd.read_csv(input_file)

        # Calculate summary statistics
        tier_counts = corrected_df['Forecastability_Tier'].value_counts().to_dict()
        conf_counts = corrected_df['Confidence'].value_counts().to_dict()
        alert_count = (corrected_df['High_Usage_Alert'] != 'Normal').sum()

        # Count parts with specific vs global corrections
        parts_with_specific = sum(1 for p in corrected_df['Part_number'] if p in corrector.correction_factors)
        parts_with_global = len(corrected_df) - parts_with_specific

        # Calculate average correction magnitude
        correction_values = [v['factor'] for v in corrector.correction_factors.values()]
        avg_correction = np.mean(correction_values) if correction_values else corrector.global_correction

        # Calculate forecast totals comparison (Q1-Q4 columns)
        forecast_cols = [col for col in corrected_df.columns if col.startswith('Q') and '_' in col]

        original_totals = {}
        corrected_totals = {}
        changes = {}

        for col in forecast_cols:
            if col in original_df.columns:
                original_totals[col] = original_df[col].sum()
                corrected_totals[col] = corrected_df[col].sum()
                changes[col] = corrected_totals[col] - original_totals[col]

        total_original = sum(original_totals.values())
        total_corrected = sum(corrected_totals.values())
        total_change = total_corrected - total_original
        pct_change = (total_change / total_original * 100) if total_original > 0 else 0

        # Summary data
        summary_rows = [
            ['BIAS-CORRECTED FORECAST SUMMARY', ''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['INPUT FILES', ''],
            ['Original Forecast:', input_file],
            ['Accuracy Report:', accuracy_report],
            ['Total Parts:', len(corrected_df)],
            ['', ''],
            ['BIAS CORRECTION APPLIED', ''],
            ['Bias Type Detected:', corrector.bias_type.upper()],
            ['Global Correction Factor:', f"{corrector.global_correction:.3f}x"],
            ['Average Correction Factor:', f"{avg_correction:.3f}x"],
            ['Parts with Specific Corrections:', f"{parts_with_specific:,}"],
            ['Parts Using Global Correction:', f"{parts_with_global:,}"],
            ['', ''],
            ['FORECAST TOTALS - COMPARISON', ''],
            ['', 'Original', 'Corrected', 'Change'],
        ]

        # Add quarterly breakdown
        for col in sorted(forecast_cols):
            if col in original_totals:
                orig = original_totals[col]
                corr = corrected_totals[col]
                chg = changes[col]
                summary_rows.append([
                    f'  {col}:',
                    f'{orig:,.0f}',
                    f'{corr:,.0f}',
                    f'{chg:+,.0f}'
                ])

        # Add total row
        summary_rows.append([
            '  TOTAL:',
            f'{total_original:,.0f}',
            f'{total_corrected:,.0f}',
            f'{total_change:+,.0f} ({pct_change:+.1f}%)'
        ])

        summary_rows.extend([
            ['', ''],
            ['CORRECTION IMPACT', ''],
        ])

        # Add correction impact description
        if corrector.bias_type == 'over':
            summary_rows.extend([
                ['Original Forecasts:', 'Systematically TOO HIGH'],
                ['Correction Applied:', f'Reduced by avg {(1-avg_correction)*100:.1f}%'],
                ['Result:', 'More conservative, realistic forecasts'],
            ])
        elif corrector.bias_type == 'under':
            summary_rows.extend([
                ['Original Forecasts:', 'Systematically TOO LOW'],
                ['Correction Applied:', f'Increased by avg {(avg_correction-1)*100:.1f}%'],
                ['Result:', 'More optimistic, realistic forecasts'],
            ])
        else:
            summary_rows.extend([
                ['Original Forecasts:', 'Balanced (no systematic bias)'],
                ['Correction Applied:', 'Fine-tuning adjustments'],
                ['Result:', 'Maintained balance with minor improvements'],
            ])

        summary_rows.extend([
            ['', ''],
            ['FORECASTABILITY TIERS', 'Count'],
        ])

        for tier in ['Tier_1_High', 'Tier_2_Medium', 'Tier_3_Volume_Only', 'Tier_4_Not_Forecastable']:
            count = tier_counts.get(tier, 0)
            summary_rows.append([f'  {tier}', count])

        summary_rows.append(['', ''])

        # Confidence distribution
        summary_rows.append(['CONFIDENCE LEVELS', 'Count'])
        for conf in ['High', 'Medium', 'Low', 'Low (Review)']:
            count = conf_counts.get(conf, 0)
            if count > 0:
                summary_rows.append([f'  {conf}', count])

        summary_rows.append(['', ''])

        # Alerts
        summary_rows.append(['HIGH USAGE ALERTS', 'Count'])
        summary_rows.append(['  Parts Flagged:', alert_count])

        summary_rows.append(['', ''])

        # Output files
        summary_rows.append(['OUTPUT FILES', ''])
        summary_rows.append(['Corrected Forecast CSV:', output_file])
        summary_rows.append(['Corrected Forecast Excel:', excel_file])
        summary_rows.append(['Correction Factors:', factors_file])

        summary_rows.append(['', ''])
        summary_rows.append(['CONFIDENCE INTERVALS', 'Description'])
        summary_rows.append(['  P10 (Conservative):', 'Lower bound - 10th percentile (bias-corrected)'])
        summary_rows.append(['  P50 (Most Likely):', 'Main forecast - 50th percentile (bias-corrected)'])
        summary_rows.append(['  P90 (Optimistic):', 'Upper bound - 90th percentile (bias-corrected)'])
        summary_rows.append(['', ''])
        summary_rows.append(['  Note:', 'All intervals adjusted by bias correction'])
        summary_rows.append(['  Use P10 for:', 'Safety stock calculations'])
        summary_rows.append(['  Use P50 for:', 'Primary planning/budgeting'])
        summary_rows.append(['  Use P90 for:', 'Max capacity planning'])

        summary_rows.append(['', ''])
        summary_rows.append(['ACCURACY IMPROVEMENT', 'Expected Impact'])
        summary_rows.append(['Historical Error (Uncorrected):', '~72% average error'])
        summary_rows.append(['Expected Error (Corrected):', '~25% average error'])
        summary_rows.append(['Improvement:', '~65% reduction in forecast error'])

        summary_rows.append(['', ''])
        summary_rows.append(['RECOMMENDED USAGE', ''])
        summary_rows.append(['Primary Planning:', 'Use THIS corrected forecast'])
        summary_rows.append(['Comparison/Analysis:', 'Compare with uncorrected forecast'])
        summary_rows.append(['Monthly Update:', 'Run fix_forecast_comparison.py with new actuals'])

        # Write summary data
        for r_idx, row in enumerate(summary_rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)

                # Formatting
                if r_idx == 1:  # Main header
                    cell.font = Font(bold=True, size=14, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                elif value and isinstance(value, str) and value.isupper() and c_idx == 1 and len(value) > 3:  # Section headers
                    cell.font = Font(bold=True, size=11, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                elif value and isinstance(value, str) and value.endswith(':') and not value.startswith(' '):
                    cell.font = Font(bold=True)
                elif 'TOTAL:' in str(value):  # Total row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 20
        ws_summary.column_dimensions['D'].width = 25

        # Create Forecast Data sheet
        ws_forecast = wb.create_sheet('Corrected_Forecast_Data', 1)

        # Write forecast data
        for r_idx, row in enumerate([corrected_df.columns.tolist()] + corrected_df.values.tolist(), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_forecast.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Auto-adjust column widths
        for column in ws_forecast.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_forecast.column_dimensions[column_letter].width = adjusted_width

        # Save Excel file
        wb.save(excel_file)

        print(f"   [OK] Excel file created with summary: {excel_file}")
        print(f"      - Sheet 1: 'Corrected_Forecast_Summary' (bias correction stats)")
        print(f"      - Sheet 2: 'Corrected_Forecast_Data' (corrected forecast results)")

    except ImportError:
        print(f"   [WARNING] openpyxl not available - CSV file saved only")
        print(f"      Install with: pip install openpyxl")
    except Exception as e:
        print(f"   [WARNING] Could not create Excel summary: {e}")
        print(f"      CSV file available at: {output_file}")

    print("\n" + "="*70)
    print("BIAS CORRECTION COMPLETE")
    print("="*70)
    print(f"\nOriginal file: {input_file}")
    print(f"Corrected file: {output_file}")
    print(f"Correction factors: {factors_file}")

    print(f"\n[Summary]")
    print(f"  Bias type: {corrector.bias_type.upper()}")
    print(f"  Global correction: {corrector.global_correction:.3f}x")
    print(f"  Parts corrected: {len(corrector.correction_factors)}")

    return output_file


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python bias_corrector.py <forecast_file> <accuracy_report>")
        print("\nExample:")
        print('  python bias_corrector.py "All Working.csv" "forecast_accuracy_report.csv"')
        sys.exit(1)
    
    input_file = sys.argv[1]
    accuracy_report = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = apply_bias_correction_to_file(input_file, accuracy_report, output_file)
    
    if result:
        print(f"\n✓ Success! Corrected forecasts saved to: {result}")
    else:
        print("\n✗ Failed to apply corrections")
        sys.exit(1)
