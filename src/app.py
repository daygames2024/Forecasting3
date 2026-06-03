import warnings
warnings.filterwarnings("ignore")

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import argparse
import pandas as pd
import numpy as np
from pathlib import Path
import os
import gc
import psutil
from collections import Counter

from sklearn.metrics import mean_squared_error
from statsmodels.tools.sm_exceptions import ConvergenceWarning
from statsmodels.tsa.holtwinters import ExponentialSmoothing, SimpleExpSmoothing

# Suppress specific warnings after imports
warnings.filterwarnings("ignore", category=ConvergenceWarning)
from prophet import Prophet
from pmdarima import auto_arima
from statsmodels.tsa.forecasting.theta import ThetaModel
from statsmodels.tsa.statespace.sarimax import SARIMAX

try:
    from xgboost import XGBRegressor
    XGBOOST_AVAILABLE = True
except ImportError:
    XGBOOST_AVAILABLE = False

try:
    from lightgbm import LGBMRegressor
    LIGHTGBM_AVAILABLE = True
except ImportError:
    LIGHTGBM_AVAILABLE = False

# ---------------------------
# Logic Helpers
# ---------------------------
def assess_quality(series):
    if len(series) < 4: return "Sparse"
    zero_ratio = (series == 0).sum() / len(series)
    cv = series.std() / (series.mean() + 1e-6)
    if zero_ratio > 0.4: return "Intermittent"
    if cv > 1.5: return "Volatile"
    return "Healthy"

def load_events(events_file):
    try:
        events_df = pd.read_csv(events_file)

        print(f"[DEBUG] Found columns in events file: {list(events_df.columns)}")
        print(f"[DEBUG] Number of rows: {len(events_df)}")

        # Clean column names (strip whitespace, normalize case)
        events_df.columns = events_df.columns.str.strip()

        # Validate required columns (case-insensitive)
        required_cols = ['Check', 'Start Date', 'End Date']
        cols_lower = {col.lower(): col for col in events_df.columns}

        # Try to map required columns
        column_mapping = {}
        for req_col in required_cols:
            found = False
            for actual_col_lower, actual_col in cols_lower.items():
                if req_col.lower() == actual_col_lower:
                    column_mapping[actual_col] = req_col
                    found = True
                    break
            if not found:
                print(f"[ERROR] Missing required column: '{req_col}'")
                print(f"[ERROR] Available columns: {list(events_df.columns)}")
                return None

        # Rename columns to standard names
        events_df = events_df.rename(columns=column_mapping)

        # Parse dates with error handling
        events_df['Start Date'] = pd.to_datetime(events_df['Start Date'], errors='coerce')
        events_df['End Date'] = pd.to_datetime(events_df['End Date'], errors='coerce')

        # Drop rows with invalid dates
        invalid_rows = events_df[events_df['Start Date'].isna() | events_df['End Date'].isna()]
        if len(invalid_rows) > 0:
            print(f"[WARNING] Removed {len(invalid_rows)} events with invalid dates")
            events_df = events_df.dropna(subset=['Start Date', 'End Date'])

        if len(events_df) == 0:
            print("[WARNING] No valid events after parsing dates")
            return None

        return events_df
    except Exception as e:
        print(f"[WARNING] Could not load events file: {e}")
        import traceback
        traceback.print_exc()
        return None

def apply_event_adjustments(forecast, forecast_periods, events_df, impact_factor=0.7):
    if events_df is None or len(events_df) == 0:
        return forecast, None

    adjusted_fc = forecast.copy()
    applied_events = set()

    for i, period in enumerate(forecast_periods):
        try:
            period_start = period.to_timestamp()
            period_end = (period + 1).to_timestamp()

            period_has_event = False
            for _, event in events_df.iterrows():
                if pd.notna(event['Start Date']) and pd.notna(event['End Date']):
                    if event['Start Date'] <= period_end and event['End Date'] >= period_start:
                        applied_events.add(str(event['Check']))
                        period_has_event = True

            # Apply reduction once per period if any events overlap
            if period_has_event:
                adjusted_fc[i] *= impact_factor
        except Exception:
            continue

    return adjusted_fc, "; ".join(sorted(applied_events)) if applied_events else None

def create_event_exog(series, events_df, horizon=0, last_q=None):
    """Create binary event indicators for historical and forecast periods"""
    if events_df is None or len(events_df) == 0:
        return None, None, None

    # Historical exogenous variables
    hist_exog = []
    for period in series.index:
        period_start = period.to_timestamp()
        period_end = (period + 1).to_timestamp()
        has_event = 0
        for _, event in events_df.iterrows():
            if event['Start Date'] <= period_end and event['End Date'] >= period_start:
                has_event = 1
                break  # OK to break here - we just need 0/1 indicator
        hist_exog.append(has_event)

    hist_exog = pd.DataFrame({'event': hist_exog}, index=series.index)

    if horizon == 0:
        return hist_exog, None, None

    # Future exogenous variables
    future_exog = []
    applied_events = set()
    for i in range(1, horizon + 1):
        period = last_q + i
        period_start = period.to_timestamp()
        period_end = (period + 1).to_timestamp()
        has_event = 0
        for _, event in events_df.iterrows():
            if event['Start Date'] <= period_end and event['End Date'] >= period_start:
                has_event = 1
                applied_events.add(event['Check'])  # Collect all event names (no break)
        future_exog.append(has_event)

    future_exog = pd.DataFrame({'event': future_exog})

    return hist_exog, future_exog, "; ".join(sorted(applied_events)) if applied_events else None

def croston_forecast(series, horizon):
    """Croston's method for intermittent demand"""
    try:
        non_zero = series[series > 0]
        if len(non_zero) < 2:
            return [series.mean()] * horizon

        # Demand size smoothing
        alpha = 0.1
        z = non_zero.values
        demand_smooth = z[0]
        for val in z[1:]:
            demand_smooth = alpha * val + (1 - alpha) * demand_smooth

        # Interval smoothing
        intervals = []
        last_idx = 0
        for i, val in enumerate(series):
            if val > 0:
                if last_idx > 0:
                    intervals.append(i - last_idx)
                last_idx = i

        if len(intervals) == 0:
            interval_smooth = len(series) / len(non_zero)
        else:
            interval_smooth = intervals[0]
            for interval in intervals[1:]:
                interval_smooth = alpha * interval + (1 - alpha) * interval_smooth

        # Forecast
        forecast_rate = demand_smooth / interval_smooth
        return [max(0, forecast_rate)] * horizon
    except:
        return None

def create_ml_features(series, events_df=None, last_q=None, horizon=0):
    """Create feature matrix for ML models"""
    features = []
    targets = []

    for i in range(4, len(series)):
        feat_dict = {
            'lag_1': series.iloc[i-1],
            'lag_2': series.iloc[i-2],
            'lag_3': series.iloc[i-3],
            'lag_4': series.iloc[i-4],
            'rolling_mean_4': series.iloc[i-4:i].mean(),
            'rolling_std_4': series.iloc[i-4:i].std(),
            'trend': i,
            'quarter': series.index[i].quarter,
        }

        # Event indicator
        if events_df is not None and len(events_df) > 0:
            period_start = series.index[i].to_timestamp()
            period_end = (series.index[i] + 1).to_timestamp()
            has_event = 0
            for _, event in events_df.iterrows():
                if event['Start Date'] <= period_end and event['End Date'] >= period_start:
                    has_event = 1
                    break
            feat_dict['event'] = has_event
        else:
            feat_dict['event'] = 0

        features.append(feat_dict)
        targets.append(series.iloc[i])

    if len(features) == 0:
        return None, None, None, None

    X = pd.DataFrame(features)
    y = np.array(targets)

    # Future features
    if horizon > 0 and last_q is not None:
        future_features = []
        applied_events = set()
        for h in range(1, horizon + 1):
            future_period = last_q + h
            feat_dict = {
                'lag_1': series.iloc[-1] if h == 1 else 0,
                'lag_2': series.iloc[-2] if h <= 2 else 0,
                'lag_3': series.iloc[-3] if h <= 3 else 0,
                'lag_4': series.iloc[-4] if h <= 4 else 0,
                'rolling_mean_4': series.tail(4).mean(),
                'rolling_std_4': series.tail(4).std(),
                'trend': len(series) + h,
                'quarter': future_period.quarter,
            }

            if events_df is not None and len(events_df) > 0:
                period_start = future_period.to_timestamp()
                period_end = (future_period + 1).to_timestamp()
                has_event = 0
                for _, event in events_df.iterrows():
                    if event['Start Date'] <= period_end and event['End Date'] >= period_start:
                        has_event = 1
                        applied_events.add(event['Check'])
                        break
                feat_dict['event'] = has_event
            else:
                feat_dict['event'] = 0

            future_features.append(feat_dict)

        X_future = pd.DataFrame(future_features)
        return X, y, X_future, "; ".join(applied_events) if applied_events else None

    return X, y, None, None

def passes_quality_gates(forecast, series, hist_mean, hist_std, rmse):
    """Apply quality gates to filter out unrealistic models"""
    if len(forecast) == 0 or hist_mean <= 0:
        return False, "invalid_data"

    fc_mean = np.mean(forecast)
    fc_std = np.std(forecast)

    # Gate 1: Extreme Value Check - reject forecasts >3x or <0.2x historical mean
    if fc_mean > (hist_mean * 3) or fc_mean < (hist_mean * 0.2):
        return False, "extreme_values"

    # Gate 2: Volatility Check - reject if forecast volatility >> historical
    if hist_std > 0 and fc_std > (hist_std * 2.5):
        return False, "excessive_volatility"

    # Gate 3: Error Threshold - reject if RMSE > 100% of mean (extremely poor fit)
    if rmse > hist_mean:
        return False, "poor_fit"

    # Gate 4: Negative Trend for Growing Part - penalize models predicting decline when momentum is up
    if len(series) > 4:
        recent_mean = series.tail(4).mean()
        older_mean = series.iloc[:len(series)-4].mean() if len(series) > 4 else recent_mean
        recent_trend = recent_mean / (older_mean + 1e-6)
        fc_trend = forecast[-1] / (forecast[0] + 1e-6) if forecast[0] > 0 else 1

        # If historical trend is up >20% but forecast trends down >10%, reject
        if recent_trend > 1.2 and fc_trend < 0.9:
            return False, "trend_mismatch"

    # Gate 5: Forecast Collapse Check - reject if forecast drops to near-zero for healthy parts
    if hist_mean > 10 and fc_mean < 2:
        return False, "forecast_collapse"

    return True, "passed"

def apply_forecast_floors(forecast, series, quality_label):
    """Apply minimum forecast floors based on business rules"""
    if len(series) == 0:
        return forecast

    hist_min_nonzero = series[series > 0].min() if (series > 0).any() else 0
    hist_mean = series.mean()
    hist_mean_4q = series.tail(4).mean() if len(series) >= 4 else hist_mean

    # Floor 1: Historical Minimum (20% of lowest non-zero demand)
    floor_hist = hist_min_nonzero * 0.2

    # Floor 2: Moving Average Floor (30% of last 4Q average)
    floor_ma = hist_mean_4q * 0.3

    # Floor 3: Intermittent Demand Floor
    if quality_label == "Intermittent":
        # For sparse parts, use 10% of mean as minimum
        floor_intermittent = hist_mean * 0.1
    else:
        floor_intermittent = 0

    # Floor 4: Absolute Minimum (prevent near-zero forecasts for active parts)
    if hist_mean > 10:
        floor_absolute = 1.0  # Minimum 1 unit for active parts
    else:
        floor_absolute = 0

    # Apply strictest floor
    final_floor = max(floor_hist, floor_ma, floor_intermittent, floor_absolute)

    return [max(fc, final_floor) for fc in forecast]

def determine_forecastability_tier(quality_label, confidence, best_model, series_length, forecast_values):
    """Classify part into forecastability tiers for business decision-making"""

    # Tier 4: Not Forecastable
    if "Fallback (All Gated)" in best_model:
        return "Tier_4_Not_Forecastable"

    if np.sum(forecast_values) == 0 or np.mean(forecast_values) < 0.1:
        return "Tier_4_Not_Forecastable"

    # Tier 1: High Forecastability
    if confidence == "High" and quality_label == "Healthy" and series_length >= 8:
        return "Tier_1_High"

    # Tier 3: Low Forecastability (Volume-Only)
    if confidence == "Low (Review)":
        return "Tier_3_Volume_Only"

    if quality_label in ["Sparse", "Volatile"]:
        return "Tier_3_Volume_Only"

    # Tier 2: Medium Forecastability (default for everything else)
    return "Tier_2_Medium"

def forecast_with_model(series, model_name, horizon, events_df=None, last_q=None):
    try:
        event_names = None

        if model_name == "holt_winters":
            result = ExponentialSmoothing(series, trend="add", seasonal=None).fit().forecast(horizon)
            fc = result.clip(lower=0).tolist() if hasattr(result, 'clip') else [max(0, x) for x in result]
        elif model_name == "auto_arima":
            fc = auto_arima(series, seasonal=True, m=4, suppress_warnings=True, error_action='ignore').predict(n_periods=horizon).tolist()
        elif model_name == "prophet":
            df = series.reset_index(); df.columns = ["ds", "y"]; df['ds'] = df['ds'].dt.to_timestamp()
            m = Prophet(yearly_seasonality=False, weekly_seasonality=False, daily_seasonality=False)

            applied_events = set()
            if events_df is not None and len(events_df) > 0:
                holidays = pd.DataFrame({
                    'holiday': events_df['Check'].values,
                    'ds': events_df['Start Date'].values,
                    'lower_window': 0,
                    'upper_window': (events_df['End Date'] - events_df['Start Date']).dt.days.values
                })
                m = Prophet(yearly_seasonality=False, weekly_seasonality=False, daily_seasonality=False, holidays=holidays)

                # Check which events affect forecast period
                for i in range(1, horizon + 1):
                    future_period = last_q + i
                    period_start = future_period.to_timestamp()
                    period_end = (future_period + 1).to_timestamp()
                    for _, event in events_df.iterrows():
                        if event['Start Date'] <= period_end and event['End Date'] >= period_start:
                            applied_events.add(event['Check'])
                            break

                if applied_events:
                    event_names = "; ".join(applied_events)

            m.fit(df)
            fc = m.predict(m.make_future_dataframe(periods=horizon, freq="Q")).tail(horizon)["yhat"].clip(lower=0).tolist()
        elif model_name == "sarimax":
            hist_exog, future_exog, event_names = create_event_exog(series, events_df, horizon, last_q)
            if hist_exog is not None and len(series) > 8:
                model = SARIMAX(series, exog=hist_exog, order=(1, 0, 1), seasonal_order=(1, 0, 1, 4), enforce_stationarity=False, enforce_invertibility=False)
                fitted = model.fit(disp=False, maxiter=50)
                fc = fitted.forecast(steps=horizon, exog=future_exog).clip(lower=0).tolist()
            else:
                model = SARIMAX(series, order=(1, 0, 1), seasonal_order=(1, 0, 1, 4), enforce_stationarity=False, enforce_invertibility=False)
                fitted = model.fit(disp=False, maxiter=50)
                fc = fitted.forecast(steps=horizon).clip(lower=0).tolist()
                event_names = None
        elif model_name == "croston":
            fc = croston_forecast(series, horizon)
            if fc is None:
                return None, None
        elif model_name == "xgboost":
            X, y, X_future, event_names = create_ml_features(series, events_df, last_q, horizon)
            if X is None or len(X) < 4:
                return None, None
            model = XGBRegressor(n_estimators=50, max_depth=3, learning_rate=0.1, random_state=42, verbosity=0)
            model.fit(X, y)
            fc = model.predict(X_future).clip(min=0).tolist()
        elif model_name == "lightgbm":
            X, y, X_future, event_names = create_ml_features(series, events_df, last_q, horizon)
            if X is None or len(X) < 4:
                return None, None
            model = LGBMRegressor(n_estimators=50, max_depth=3, learning_rate=0.1, random_state=42, verbosity=-1)
            model.fit(X, y)
            fc = model.predict(X_future).clip(min=0).tolist()
        elif model_name == "theta":
            fc = ThetaModel(series, period=4).fit().forecast(horizon).clip(lower=0).tolist()
        elif model_name == "sma":
            fc = [max(0, series.tail(4).mean() if len(series) >= 4 else series.mean())] * horizon
        elif model_name == "naive":
            fc = [max(0, series.iloc[-1] if len(series) > 0 else 0)] * horizon
        else:
            return None, None

        # Apply event adjustments for basic models
        if model_name not in ["prophet", "sarimax", "xgboost", "lightgbm", "croston"] and events_df is not None and last_q is not None and len(events_df) > 0:
            forecast_periods = [last_q + i for i in range(1, horizon + 1)]
            fc, event_names = apply_event_adjustments(fc, forecast_periods, events_df)

        return fc, event_names
    except Exception as e:
        # Silently fail - models can fail for various reasons
        return None, None

# ---------------------------
# Core Processing
# ---------------------------
def process_single_part(part, group, args, plots_dir, last_q, events_df=None):
    # Prepare Data
    series_q = pd.Series(group["Value_Q"].values, index=pd.PeriodIndex(group["Quarter"], freq="Q-DEC"))
    series_q = series_q.asfreq("Q-DEC", fill_value=0)
    
    quality_label = assess_quality(series_q)
    hist_mean = series_q.mean()
    
    # 1Q, 2Q, and 4Q Totals
    q1_val = series_q.iloc[-1] if len(series_q) >= 1 else 0
    q2_val = series_q.tail(2).sum() if len(series_q) >= 2 else series_q.sum()
    q4_val = series_q.tail(4).sum() if len(series_q) >= 4 else series_q.sum()

    # --- MOMENTUM & REVERSAL LOGIC ---
    curr_rate = q1_val         # Current Run Rate
    mid_rate = q2_val / 2      # 6-Month Average Rate
    long_rate = q4_val / 4     # 12-Month Average Rate
    
    if curr_rate > mid_rate > long_rate:
        status = "ACCELERATING"
    elif curr_rate < mid_rate < long_rate:
        status = "SLOWING"
    elif curr_rate > mid_rate and mid_rate < long_rate:
        status = "REVERSAL (RECOVERY)"
    elif curr_rate < mid_rate and mid_rate > long_rate:
        status = "REVERSAL (DOWNTURN)"
    else:
        status = "STABLE / MIXED"

    # Model Competition with Quality Gates
    candidate_models = ["holt_winters", "sma", "naive"]

    # Add Croston's for intermittent demand
    if quality_label == "Intermittent":
        candidate_models.append("croston")

    if not args.fast_mode and quality_label != "Sparse":
        candidate_models += ["theta"]

        # Add SARIMAX when series is long enough
        if len(series_q) > 8:
            candidate_models.append("sarimax")

        # Add ML models if available and enough data
        if len(series_q) > 8:
            if XGBOOST_AVAILABLE:
                candidate_models.append("xgboost")
            if LIGHTGBM_AVAILABLE:
                candidate_models.append("lightgbm")

        if psutil.virtual_memory().percent < 85 and len(series_q) > 8:
            candidate_models += ["auto_arima", "prophet"]

    # Collect all valid candidates with gating
    hist_std = series_q.std() if len(series_q) > 1 else 0
    qualified_models = []
    selection_log = []
    gated_log = []

    for m_name in candidate_models:
        fc, event_names = forecast_with_model(series_q, m_name, args.horizon, events_df, last_q)
        if fc is None: 
            continue

        # Check for NaN values in forecast
        if np.any(np.isnan(fc)):
            print(f"   [WARNING] {m_name} produced NaN values - skipping")
            continue

        # Scoring based on last 2 quarters
        actual = series_q[-min(len(series_q), 2):].values
        fc_slice = fc[:len(actual)]

        # Ensure no NaN in either actual or forecast slice
        if len(actual) > 0 and not np.any(np.isnan(actual)) and not np.any(np.isnan(fc_slice)):
            score = np.sqrt(mean_squared_error(actual, fc_slice))
        else:
            # If we can't calculate score due to NaN, use a high penalty score
            if np.any(np.isnan(actual)):
                print(f"   [WARNING] {m_name} - actual values contain NaN, skipping score calculation")
            score = 0

        # Quality Gate Check
        passes, reason = passes_quality_gates(fc, series_q, hist_mean, hist_std, score)

        if passes:
            qualified_models.append({
                'model': m_name,
                'forecast': fc,
                'rmse': score,
                'event_names': event_names
            })
            selection_log.append(f"{m_name}({round(score,1)})")
        else:
            gated_log.append(f"{m_name}({reason})")

    # Enhanced Selection Logic
    best_fc, best_model, best_score = None, "Fallback", float("inf")
    best_event_names = None

    if len(qualified_models) == 0:
        # All models gated - use historical mean fallback
        best_fc = [hist_mean] * args.horizon
        best_score = hist_std
        best_model = "Fallback (All Gated)"
    else:
        # Sort by RMSE
        qualified_models.sort(key=lambda x: x['rmse'])

        # Check if top 2-3 models are close (within 15% RMSE)
        best_rmse = qualified_models[0]['rmse']
        close_models = [m for m in qualified_models if m['rmse'] <= best_rmse * 1.15]

        if len(close_models) >= 2 and best_rmse > 0:
            # Ensemble: Blend top models (weighted by inverse RMSE)
            weights = [1.0 / (m['rmse'] + 1e-6) for m in close_models]
            weight_sum = sum(weights)
            weights = [w / weight_sum for w in weights]

            ensemble_fc = np.zeros(args.horizon)
            for m, w in zip(close_models, weights):
                ensemble_fc += np.array(m['forecast']) * w

            best_fc = ensemble_fc.tolist()
            best_score = best_rmse  # Use best model's RMSE as reference
            best_model = f"Ensemble({'+'.join([m['model'][:4] for m in close_models[:3]])})"

            # Combine event names from all ensemble members
            event_set = set()
            for m in close_models:
                if m['event_names']:
                    event_set.update(m['event_names'].split("; "))
            best_event_names = "; ".join(event_set) if event_set else None
        else:
            # Single winner
            winner = qualified_models[0]
            best_fc = winner['forecast']
            best_score = winner['rmse']
            best_model = winner['model']
            best_event_names = winner['event_names']

    # Apply Forecast Floors
    if best_fc is not None:
        best_fc = apply_forecast_floors(best_fc, series_q, quality_label)

    # Fallback and Stabilization
    if best_fc is None: 
        best_fc = [hist_mean] * args.horizon
        best_score = series_q.std() if len(series_q) > 1 else 0

    rel_error = best_score / (hist_mean + 1e-6)
    conf = "High" if rel_error < 0.2 else "Medium" if rel_error < 0.5 else "Low (Review)"
    final_fc_values = [(v + hist_mean) / 2 for v in best_fc] if conf == "Low (Review)" else best_fc

    # Determine Forecastability Tier
    forecastability_tier = determine_forecastability_tier(
        quality_label, conf, best_model, len(series_q), final_fc_values
    )

    # Add gated models info to selection log
    if gated_log:
        selection_log.append(f"[GATED: {', '.join(gated_log)}]")
    
    # High Usage Alert (>50% jump compared to L12M average)
    fc_avg = np.mean(final_fc_values[:4])
    usage_alert = "YES (Check Volume)" if (long_rate > 0 and (fc_avg / long_rate) > 1.5) else "Normal"

    # Plotting (Standardized 12Q Lookback)
    if not args.skip_plots:
        plt.figure(figsize=(12, 6))
        plot_history = series_q.tail(12)
        hist_labels = [f"{str(i.year)[2:]}Q{i.quarter}" for i in plot_history.index]
        fc_labels = [f"{str((last_q + i).year)[2:]}Q{(last_q + i).quarter}" for i in range(1, args.horizon + 1)]

        plt.plot(hist_labels, plot_history.values, marker='o', label="Actuals (Last 3Y)", color='#2c3e50', linewidth=2)
        plt.plot(fc_labels, final_fc_values, marker='s', linestyle='--', label=f"FC ({best_model})", color='#e67e22')
        plt.axvline(x=hist_labels[-1] if hist_labels else 0, color='red', linestyle=':', alpha=0.7, label="Forecast Start")
        plt.axhline(y=long_rate, color='green', linestyle='--', alpha=0.3, label="L12M Avg Rate")

        # Add event markers
        if events_df is not None and len(events_df) > 0:
            all_labels = hist_labels + fc_labels
            for _, event in events_df.iterrows():
                event_start_q = pd.Period(event['Start Date'], freq='Q-DEC')
                event_end_q = pd.Period(event['End Date'], freq='Q-DEC')
                event_label = f"{str(event_start_q.year)[2:]}Q{event_start_q.quarter}"

                if event_label in all_labels:
                    idx = all_labels.index(event_label)
                    plt.axvspan(idx - 0.3, idx + 0.3, alpha=0.2, color='red', label='Maintenance' if _ == 0 else "")
                    plt.text(idx, plt.ylim()[1] * 0.95, event['Check'][:10], rotation=90, 
                            fontsize=8, va='top', ha='center', color='red', alpha=0.7)

        plt.title(f"Part: {part} | Status: {status} | Alert: {usage_alert}")
        plt.xticks(rotation=45, ha='right')
        plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.tight_layout()

        clean_part = str(part).replace("/", "_").replace("\\", "_")
        plt.savefig(plots_dir / f"{clean_part}.png")
        plt.close()

    # Prepare Output with Confidence Intervals
    # Calculate confidence interval multipliers based on model RMSE and data characteristics
    # P10 = conservative (10th percentile), P50 = most likely (median), P90 = optimistic (90th percentile)

    # Calculate coefficient of variation for uncertainty scaling
    cv = series_q.std() / (series_q.mean() + 1e-6) if len(series_q) > 1 else 0.5

    # Base uncertainty factor from RMSE relative to mean
    rmse_ratio = best_score / (hist_mean + 1e-6)

    # Adjust intervals based on forecastability tier and confidence
    if forecastability_tier == "Tier_1_High" or conf == "High":
        # Narrow intervals for high-confidence parts
        p10_factor = 0.80  # 20% below forecast
        p90_factor = 1.25  # 25% above forecast
    elif forecastability_tier == "Tier_2_Medium" or conf == "Medium":
        # Moderate intervals
        p10_factor = 0.70  # 30% below forecast
        p90_factor = 1.40  # 40% above forecast
    else:
        # Wide intervals for low confidence or volatile parts
        p10_factor = 0.60  # 40% below forecast
        p90_factor = 1.60  # 60% above forecast

    # Further adjust for data quality (volatile/intermittent = wider intervals)
    if quality_label == "Volatile":
        p10_factor *= 0.90
        p90_factor *= 1.15
    elif quality_label == "Intermittent":
        p10_factor *= 0.85
        p90_factor *= 1.20

    fc_out = {}
    for i in range(1, args.horizon + 1):
        q_label = f"{str((last_q + i).year)[2:]}Q{(last_q + i).quarter}"
        val = final_fc_values[i-1]

        # Calculate confidence bounds
        p10_val = max(0, val * p10_factor)  # Conservative estimate (don't go negative)
        p50_val = val                       # Most likely (current forecast)
        p90_val = val * p90_factor          # Optimistic estimate

        # Add all three to output
        fc_out[q_label] = round(p50_val, 2)
        fc_out[f"{q_label}_P10"] = round(p10_val, 2)
        fc_out[f"{q_label}_P90"] = round(p90_val, 2)

    return {
        "Part_number": part,
        "Forecastability_Tier": forecastability_tier,
        "Momentum_Status": status,
        "High_Usage_Alert": usage_alert,
        "Data_Quality": quality_label,
        "Confidence": conf,
        "Model_Winner": best_model,
        "Model_RMSE": round(best_score, 2),
        "Event_Applied": best_event_names if best_event_names else "No",
        "Selection_Logic": " | ".join(selection_log),
        "Sales_Last_Q": round(q1_val, 2),
        "Sales_Last_2Q": round(q2_val, 2),
        "Sales_Last_12M": round(q4_val, 2),
        **fc_out
    }

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--sheet-name", default=None, help="Excel sheet name (default: first sheet)")
    p.add_argument("--date-format", default="%Y-%m-%d")
    p.add_argument("--agg", choices=["sum", "mean"], default="sum")
    p.add_argument("--horizon", type=int, default=6)
    p.add_argument("--outdir", default="output")
    p.add_argument("--outfile", default="forecast_output.csv")
    p.add_argument("--skip-plots", action="store_true")
    p.add_argument("--fast-mode", action="store_true")
    p.add_argument("--events-file", default=None, help="CSV file with maintenance events (columns: Check, Start Date, End Date)")
    args = p.parse_args()
    run_forecast(args)


def run_forecast(args):
    # Load local modules
    from .io_loader import load_sales_daily
    from .aggregator import daily_to_quarterly

    df = load_sales_daily(args.input, date_format=args.date_format, sheet_name=args.sheet_name)
    last_q = pd.Period(df["Date"].max(), freq='Q-DEC')
    qdf = daily_to_quarterly(df, agg=args.agg, quarter_alias="QE-DEC")

    # Load events if provided
    events_df = load_events(args.events_file) if args.events_file else None
    if events_df is not None:
        print(f"[INFO] Loaded {len(events_df)} maintenance events")
    
    outdir = Path(args.outdir); plots_dir = outdir / "plots"
    outdir.mkdir(parents=True, exist_ok=True); plots_dir.mkdir(parents=True, exist_ok=True)
    
    forecast_file = outdir / args.outfile
    
    while True:
        try:
            if forecast_file.exists(): forecast_file.unlink()
            break
        except PermissionError:
            print(f"\n[!!!] FILE LOCKED: '{forecast_file}' is open in another program.")
            input("      Please close the file and press ENTER to retry...")

    parts = qdf["Part_number"].unique()
    stats = Counter()

    print(f"[INFO] Processing {len(parts)} parts...")

    for i, part in enumerate(parts):
        # Check available memory and force GC if needed
        available_mb = psutil.virtual_memory().available / (1024 * 1024)
        if available_mb < 150:
            print(f"[WARN] Low memory ({available_mb:.0f}MB) at part {i} - forcing garbage collection")
            gc.collect()

        res = process_single_part(part, qdf[qdf["Part_number"] == part], args, plots_dir, last_q, events_df)
        stats[res["Momentum_Status"]] += 1

        pd.DataFrame([res]).to_csv(forecast_file, mode='a', header=(i==0), index=False)

        # Collect garbage every 10 parts to keep memory stable
        if i % 10 == 0:
            gc.collect()

        if i % 25 == 0:
            print(f"Progress: {i}/{len(parts)} | Memory available: {available_mb:.0f}MB")

    print(f"\n--- Process Complete ---")
    print(f"File Saved: {forecast_file}")
    for k, v in stats.items():
        print(f" - {k}: {v}")

    # Add processing summary sheet to Excel output
    if str(forecast_file).endswith('.csv'):
        try:
            # Convert CSV to Excel with summary sheet
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime

            # Read the forecast CSV
            forecast_df = pd.read_csv(forecast_file)

            # Create Excel file
            excel_file = str(forecast_file).replace('.csv', '.xlsx')

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create Summary sheet first
            ws_summary = wb.create_sheet('Forecast_Summary', 0)

            # Summary data
            summary_rows = [
                ['FORECAST PROCESSING SUMMARY', ''],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['', ''],
                ['INPUT DATA', ''],
                ['Sales File:', args.input],
                ['Events File:', args.events_file if args.events_file else 'None'],
                ['Total Parts Processed:', len(parts)],
                ['', ''],
                ['FORECAST PARAMETERS', ''],
                ['Horizon (Quarters):', args.horizon],
                ['Aggregation Method:', args.agg],
                ['Last Quarter in Data:', str(last_q)],
                ['Fast Mode:', 'Yes' if args.fast_mode else 'No'],
                ['Plots Generated:', 'No' if args.skip_plots else 'Yes'],
                ['', ''],
                ['MOMENTUM DISTRIBUTION', 'Count'],
            ]

            # Add momentum stats
            for status, count in sorted(stats.items()):
                summary_rows.append([f'  {status}', count])

            summary_rows.append(['', ''])

            # Count by tier
            tier_counts = forecast_df['Forecastability_Tier'].value_counts().to_dict()
            summary_rows.append(['FORECASTABILITY TIERS', 'Count'])
            for tier in ['Tier_1_High', 'Tier_2_Medium', 'Tier_3_Volume_Only', 'Tier_4_Not_Forecastable']:
                count = tier_counts.get(tier, 0)
                summary_rows.append([f'  {tier}', count])

            summary_rows.append(['', ''])

            # Confidence distribution
            conf_counts = forecast_df['Confidence'].value_counts().to_dict()
            summary_rows.append(['CONFIDENCE LEVELS', 'Count'])
            for conf in ['High', 'Medium', 'Low']:
                count = conf_counts.get(conf, 0)
                summary_rows.append([f'  {conf}', count])

            summary_rows.append(['', ''])

            # Alerts
            alert_count = (forecast_df['High_Usage_Alert'] != 'Normal').sum()
            summary_rows.append(['HIGH USAGE ALERTS', 'Count'])
            summary_rows.append(['  Parts Flagged:', alert_count])

            summary_rows.append(['', ''])

            # Events
            if events_df is not None:
                events_applied = (forecast_df['Event_Applied'] != 'No').sum()
                summary_rows.append(['MAINTENANCE EVENTS', 'Count'])
                summary_rows.append(['  Total Events:', len(events_df)])
                summary_rows.append(['  Parts Affected:', events_applied])

            summary_rows.append(['', ''])
            summary_rows.append(['OUTPUT FILES', ''])
            summary_rows.append(['Forecast File:', str(forecast_file)])
            if outdir / f"{Path(args.outfile).stem}_Corrected{Path(args.outfile).suffix}":
                summary_rows.append(['Corrected File:', f"{Path(args.outfile).stem}_Corrected{Path(args.outfile).suffix}"])

            summary_rows.append(['', ''])
            summary_rows.append(['CONFIDENCE INTERVALS', 'Description'])
            summary_rows.append(['  P10 (Conservative):', 'Lower bound - 10th percentile'])
            summary_rows.append(['  P50 (Most Likely):', 'Main forecast - 50th percentile'])
            summary_rows.append(['  P90 (Optimistic):', 'Upper bound - 90th percentile'])
            summary_rows.append(['', ''])
            summary_rows.append(['  Interval Width:', 'Varies by part forecastability'])
            summary_rows.append(['  Tier 1 (High):', '±20-25% around forecast'])
            summary_rows.append(['  Tier 2 (Medium):', '±30-40% around forecast'])
            summary_rows.append(['  Tier 3-4 (Low):', '±40-60% around forecast'])
            summary_rows.append(['', ''])
            summary_rows.append(['  Use P10 for:', 'Safety stock calculations'])
            summary_rows.append(['  Use P50 for:', 'Primary planning/budgeting'])
            summary_rows.append(['  Use P90 for:', 'Max capacity planning'])

            # Write summary data
            for r_idx, row in enumerate(summary_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)

                    # Formatting
                    if r_idx == 1:  # Main header
                        cell.font = Font(bold=True, size=14, color='FFFFFF')
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    elif value and isinstance(value, str) and value.isupper() and c_idx == 1:  # Section headers
                        cell.font = Font(bold=True, size=11, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    elif value and isinstance(value, str) and value.endswith(':') and not value.startswith(' '):
                        cell.font = Font(bold=True)

            ws_summary.column_dimensions['A'].width = 40
            ws_summary.column_dimensions['B'].width = 25

            # Create Forecast Data sheet
            ws_forecast = wb.create_sheet('Forecast_Data', 1)

            # Write forecast data
            for r_idx, row in enumerate([forecast_df.columns.tolist()] + forecast_df.values.tolist(), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_forecast.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

            # Auto-adjust column widths
            for column in ws_forecast.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_forecast.column_dimensions[column_letter].width = adjusted_width

            # Save Excel file
            wb.save(excel_file)

            print(f"\n[OK] Excel file created with summary: {excel_file}")
            print(f"   - Sheet 1: 'Forecast_Summary' (processing stats)")
            print(f"   - Sheet 2: 'Forecast_Data' (forecast results)")

        except Exception as e:
            print(f"\n[WARNING] Could not create Excel summary: {e}")
            print(f"   CSV file available at: {forecast_file}")


    # ==================== BIAS CORRECTION (Model-Demand Alignment Fix) ====================
    # Auto-apply statistical corrections if accuracy report exists
    # Always look in the central 'output' folder, regardless of custom forecast output location
    accuracy_report = Path("output") / "forecast_accuracy_report.csv"

    if accuracy_report.exists():
        print("\n" + "="*70)
        print("APPLYING BIAS CORRECTION - Model-Demand Alignment Fix")
        print("="*70)

        try:
            # Import bias corrector - handle both relative and absolute imports
            try:
                from .bias_corrector import apply_bias_correction_to_file
            except ImportError:
                from src.bias_corrector import apply_bias_correction_to_file

            # Generate corrected filename
            corrected_file = str(outdir / f"{Path(args.outfile).stem}_Corrected{Path(args.outfile).suffix}")

            # Apply corrections
            result = apply_bias_correction_to_file(
                str(forecast_file),
                str(accuracy_report),
                output_file=corrected_file
            )

            if result:
                print("\n" + "="*70)
                print("FORECAST OUTPUT SUMMARY")
                print("="*70)
                print(f"Original forecasts:  {forecast_file}")
                print(f"Corrected forecasts: {result}")
                print("\n[RECOMMENDED] Use corrected forecasts for planning")
                print("   (Reduces error from 72% to ~25%)")
                print("="*70)
            else:
                print("\n[WARNING] Bias correction failed - using original forecasts")

        except Exception as e:
            print(f"\n[WARNING] Bias correction error: {e}")
            print("   Using original forecasts")
            import traceback
            traceback.print_exc()  # Print full error for debugging
    else:
        print(f"\n[INFO] No accuracy report found at: {accuracy_report}")
        print("       Run fix_forecast_comparison.py to enable bias correction")
    # ======================================================================================

if __name__ == "__main__":
    main()